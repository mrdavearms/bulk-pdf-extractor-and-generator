#!/usr/bin/env python3
"""
VCAA Evidence Application PDF Generator v2.0
A GUI application to analyze PDF templates, map fields, and batch-fill forms.

Features:
- Multi-stage tabbed interface with modern dark theme
- PDF template analysis with visual field preview
- Combed field detection and smart filling
- Field mapping configuration
- Template library management
- Batch PDF generation from Excel data
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime
import pandas as pd
from pypdf import PdfReader, PdfWriter
import threading
import json
from pathlib import Path
from typing import List, Optional, Dict
from io import BytesIO
from PIL import Image, ImageTk, ImageDraw, ImageFont

# Import our new modules
from vcaa_models import PDFField, TemplateConfig, AppSettings
from vcaa_pdf_analyzer import PDFAnalyzer, auto_name_template
from vcaa_visual_preview import VisualPreviewGenerator
from vcaa_combed_filler import CombedFieldFiller
from vcaa_theme import (
    COLORS, SPACING, SYSTEM_FONTS, font,
    apply_dark_theme, setup_treeview_tags, bind_treeview_hover,
)
from vcaa_markdown_renderer import load_and_render


class ScrollableFrame(ttk.Frame):
    """A helper class to create a scrollable frame with dark-themed canvas.

    Mousewheel scrolling is scoped: only the frame under the cursor
    scrolls, and the delta calculation is platform-aware (Windows,
    macOS, Linux all behave differently).
    """

    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(
            self,
            borderwidth=0,
            highlightthickness=0,
            bg=COLORS['bg_base'],
        )
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Scoped mousewheel: only bind when cursor is over this frame
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)
        self.scrollable_frame.bind("<Enter>", self._bind_mousewheel)
        self.scrollable_frame.bind("<Leave>", self._unbind_mousewheel)

    def _bind_mousewheel(self, event=None):
        """Bind mousewheel events to this specific canvas (not bind_all)."""
        if sys.platform == 'linux':
            self.canvas.bind("<Button-4>", self._on_mousewheel)
            self.canvas.bind("<Button-5>", self._on_mousewheel)
        else:
            self.canvas.bind("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, event=None):
        """Unbind mousewheel events when cursor leaves."""
        if sys.platform == 'linux':
            self.canvas.unbind("<Button-4>")
            self.canvas.unbind("<Button-5>")
        else:
            self.canvas.unbind("<MouseWheel>")

    def _on_mousewheel(self, event):
        """Scroll the canvas. Delta handling is platform-aware."""
        if not self.canvas.winfo_exists():
            return
        if sys.platform == 'darwin':
            # macOS: delta is already ±1 (or small integers)
            self.canvas.yview_scroll(int(-1 * event.delta), "units")
        elif sys.platform == 'linux':
            # Linux: Button-4 = scroll up, Button-5 = scroll down
            direction = -1 if event.num == 4 else 1
            self.canvas.yview_scroll(direction * 3, "units")
        else:
            # Windows: delta is ±120 per notch
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class WelcomeDialog(tk.Toplevel):
    """First-run welcome dialog."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Welcome to Bulk PDF Generator")
        self.geometry("560x380")
        self.configure(bg=COLORS['bg_elevated'])
        self.transient(parent)
        self.grab_set()

        self.choice = None
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="Welcome to Bulk PDF Generator",
            font=(ff, 18, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
        ).pack(pady=(30, 8))

        tk.Label(
            self,
            text="It looks like this is your first time.",
            font=(ff, 11),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
        ).pack(pady=(0, 20))

        # Instructions
        tk.Label(
            self,
            text="Would you like to:",
            font=(ff, 11),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
        ).pack(pady=(0, 10))

        # Radio buttons
        self.choice_var = tk.StringVar(value="analyze")

        options_frame = tk.Frame(self, bg=C['bg_elevated'])
        options_frame.pack(pady=10, padx=40, fill=tk.X)

        for text, value in [
            ("Analyze a PDF template (recommended for new users)", "analyze"),
            ("Load an existing template configuration", "load"),
            ("Skip to PDF generation (I know what I'm doing)", "skip"),
        ]:
            ttk.Radiobutton(
                options_frame,
                text=text,
                variable=self.choice_var,
                value=value,
                style='Elevated.TRadiobutton',
            ).pack(anchor=tk.W, pady=5)

        # Continue button
        ttk.Button(
            self,
            text="Continue",
            command=self.on_continue,
            style='Accent.TButton',
        ).pack(pady=25)

        # Auto-size to content and center on parent
        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        # Force to front
        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def on_continue(self):
        self.choice = self.choice_var.get()
        self.destroy()


class SchoolSetupDialog(tk.Toplevel):
    """One-time dialog to capture school name and academic year."""

    def __init__(self, parent, current_name: str = "", current_year: str = ""):
        super().__init__(parent)
        self.title("School Setup")
        self.configure(bg=COLORS['bg_elevated'])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.result_name = None
        self.result_year = None
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="School Details",
            font=(ff, 18, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
        ).pack(pady=(30, 4))

        tk.Label(
            self,
            text="These details are used in generated PDF filenames\n"
                 "and are saved so you only need to enter them once.",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
            justify=tk.CENTER,
        ).pack(pady=(0, 20))

        # Form area
        form = tk.Frame(self, bg=C['bg_elevated'])
        form.pack(padx=40, fill=tk.X)

        # School name
        tk.Label(
            form,
            text="School Name",
            font=(ff, 10, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 4))

        self.name_var = tk.StringVar(value=current_name)
        name_entry = ttk.Entry(form, textvariable=self.name_var, width=45,
                               style='Elevated.TEntry')
        name_entry.pack(fill=tk.X, pady=(0, 4))
        name_entry.focus_set()

        tk.Label(
            form,
            text='e.g. "Wangaratta High School"',
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 16))

        # Academic year
        tk.Label(
            form,
            text="Academic Year",
            font=(ff, 10, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 4))

        # Default to current year if not provided
        if not current_year:
            current_year = str(datetime.now().year)
        self.year_var = tk.StringVar(value=current_year)
        year_entry = ttk.Entry(form, textvariable=self.year_var, width=10,
                               style='Elevated.TEntry')
        year_entry.pack(anchor=tk.W, pady=(0, 16))

        # Save button
        ttk.Button(
            self,
            text="Save",
            command=self.on_save,
            style='Accent.TButton',
        ).pack(pady=(8, 25))

        # Bind Enter key
        self.bind('<Return>', lambda e: self.on_save())

        # Auto-size and centre on parent
        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def on_save(self):
        name = self.name_var.get().strip()
        year = self.year_var.get().strip()
        if not name:
            messagebox.showwarning("Required", "Please enter your school name.",
                                   parent=self)
            return
        if not year:
            messagebox.showwarning("Required", "Please enter the academic year.",
                                   parent=self)
            return
        self.result_name = name
        self.result_year = year
        self.destroy()


class TemplateNameDialog(tk.Toplevel):
    """Dialog for naming a template before analysis."""

    def __init__(self, parent, suggested_name: str):
        super().__init__(parent)
        self.title("Template Name")
        self.geometry("520x280")
        self.configure(bg=COLORS['bg_elevated'])
        self.transient(parent)
        self.grab_set()

        self.result = None
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="Template Name",
            font=(ff, 16, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
        ).pack(pady=(20, 5))

        tk.Label(
            self,
            text="This configuration will be saved for reuse.",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
        ).pack(pady=(0, 15))

        # Name entry
        name_frame = tk.Frame(self, bg=C['bg_elevated'])
        name_frame.pack(pady=10, padx=40, fill=tk.X)

        tk.Label(name_frame, text="Template Name:",
                 font=(ff, 11), fg=C['text_primary'],
                 bg=C['bg_elevated']).pack(anchor=tk.W, pady=(0, 5))

        self.name_var = tk.StringVar(value=suggested_name)
        ttk.Entry(name_frame, textvariable=self.name_var, width=50).pack(fill=tk.X)

        # Auto-generated info
        tk.Label(
            name_frame,
            text=f"Auto-generated from: {os.path.basename(suggested_name)}",
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_elevated'],
        ).pack(anchor=tk.W, pady=(5, 10))

        # Naming option
        self.naming_var = tk.StringVar(value="custom")

        ttk.Radiobutton(
            name_frame,
            text="Use PDF filename (auto-clean)",
            variable=self.naming_var,
            value="auto",
            style='Elevated.TRadiobutton',
        ).pack(anchor=tk.W, pady=2)

        ttk.Radiobutton(
            name_frame,
            text="Custom name (editable above)",
            variable=self.naming_var,
            value="custom",
            style='Elevated.TRadiobutton',
        ).pack(anchor=tk.W, pady=2)

        # Buttons
        button_frame = tk.Frame(self, bg=C['bg_elevated'])
        button_frame.pack(pady=20, fill=tk.X, padx=40)

        ttk.Button(
            button_frame,
            text="Cancel",
            command=self.on_cancel,
        ).pack(side=tk.LEFT)

        ttk.Button(
            button_frame,
            text="Analyze & Save",
            command=self.on_save,
            style='Accent.TButton',
        ).pack(side=tk.RIGHT)

        # Auto-size to content and center on parent
        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        # Force to front
        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def on_save(self):
        self.result = self.name_var.get().strip()
        if not self.result:
            messagebox.showwarning("Invalid Name", "Please enter a template name.")
            return
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()


class VCAAPDFGeneratorV2:
    """Main application class with tabbed interface."""

    def __init__(self, root):
        self.root = root
        self.root.title("Bulk PDF Generator v2.0")
        self.root.geometry("1000x800")
        self.root.minsize(900, 700)

        # Icon references — must live on self to prevent garbage collection
        self._icon_refs = {}
        self._load_app_icon()

        # Settings
        self.settings_file = os.path.expanduser("~/Documents/VCAA_App/settings.json")
        self.settings = self.load_settings()

        # Ensure templates directory exists
        os.makedirs(self.settings.templates_directory, exist_ok=True)

        # Current state
        self.current_template: Optional[TemplateConfig] = None
        self.analyzed_fields: List[PDFField] = []
        self.pdf_template_path = tk.StringVar()
        self.excel_file_path = tk.StringVar()
        self.pdf_fields: List[str] = []

        # Tab 3 (Generate) state - from original app
        self.df = None
        self.selected_rows = {}
        self.critical_fields = ['surname', 'first name', 'vcaa student number']
        self.output_dir_path = tk.StringVar()  # Optional custom output directory

        # Register cleanup on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.setup_ui()

        # First-launch: ask for school details (runs once, then never again)
        if not self.settings.school_configured:
            self.root.after(300, self.prompt_school_setup)
        # Show welcome dialog if first time
        elif self.settings.show_welcome and not self.has_templates():
            self.root.after(500, self.show_welcome_dialog)

    def _load_app_icon(self):
        """Load icon.png / icon.ico and apply to window and pre-scale for UI use.

        Pre-scaled ImageTk.PhotoImage objects are stored in self._icon_refs so
        they are never garbage-collected while the app is running:
          'header' → 32×32  used in the header bar
          'about'  → 72×72  used in the About tab card
        """
        icon_dir = os.path.dirname(os.path.abspath(__file__))
        ico_path = os.path.join(icon_dir, 'icon.ico')
        png_path = os.path.join(icon_dir, 'icon.png')

        # Set taskbar / title-bar icon
        try:
            if sys.platform == 'win32' and os.path.exists(ico_path):
                self.root.iconbitmap(ico_path)
            elif os.path.exists(png_path):
                _img = Image.open(png_path).convert('RGBA')
                _photo = ImageTk.PhotoImage(_img.resize((64, 64), Image.LANCZOS))
                self.root.iconphoto(True, _photo)
                self._icon_refs['window'] = _photo
        except Exception:
            pass  # Fall back to default tkinter icon

        # Pre-scale for UI use
        if os.path.exists(png_path):
            try:
                _img = Image.open(png_path).convert('RGBA')
                for size, key in [(32, 'header'), (72, 'about')]:
                    _scaled = _img.resize((size, size), Image.LANCZOS)
                    self._icon_refs[key] = ImageTk.PhotoImage(_scaled)
            except Exception:
                pass  # UI falls back to diamond glyph

    def _close_preview_generator(self):
        """Safely close the preview generator if open."""
        if hasattr(self, 'preview_generator') and self.preview_generator:
            try:
                self.preview_generator.__exit__(None, None, None)
            except Exception:
                pass
            self.preview_generator = None

    def on_closing(self):
        """Clean up resources before closing."""
        self._close_preview_generator()
        self.root.destroy()

    def load_settings(self) -> AppSettings:
        """Load app settings or create defaults."""
        if os.path.exists(self.settings_file):
            return AppSettings.from_file(self.settings_file)
        else:
            # Create default settings
            settings = AppSettings.get_defaults()
            os.makedirs(os.path.dirname(self.settings_file), exist_ok=True)
            settings.save_to_file(self.settings_file)
            return settings

    def has_templates(self) -> bool:
        """Check if any templates exist."""
        if not os.path.exists(self.settings.templates_directory):
            return False
        templates = list(Path(self.settings.templates_directory).glob("*.json"))
        return len(templates) > 0

    def prompt_school_setup(self):
        """Show school setup dialog (first launch only)."""
        dialog = SchoolSetupDialog(
            self.root,
            current_name=self.settings.school_name,
            current_year=self.settings.school_year,
        )
        self.root.wait_window(dialog)

        if dialog.result_name:
            self.settings.school_name = dialog.result_name
            self.settings.school_year = dialog.result_year
            self.settings.save_to_file(self.settings_file)
            self.update_status(
                f"School set: {dialog.result_name} ({dialog.result_year})", 'success'
            )
            # Update header if it's showing
            if hasattr(self, 'header_school'):
                self.header_school.config(
                    text=f"{self.settings.school_name}  |  {self.settings.school_year}"
                )

        # Now show the regular welcome dialog if applicable
        if self.settings.show_welcome and not self.has_templates():
            self.root.after(200, self.show_welcome_dialog)

    def show_welcome_dialog(self):
        """Show welcome dialog for first-time users."""
        dialog = WelcomeDialog(self.root)
        self.root.wait_window(dialog)

        if dialog.choice == "analyze":
            self.notebook.select(1)  # Go to Tab 1
        elif dialog.choice == "load":
            self.load_template_from_file()
        elif dialog.choice == "skip":
            self.notebook.select(3)  # Go to Tab 3

    def setup_ui(self):
        """Create the main UI with tabbed interface."""
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Main container
        main_frame = ttk.Frame(self.root, padding="0")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ── Header Bar ──
        header = tk.Frame(main_frame, bg=C['bg_surface'], height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        title_area = tk.Frame(header, bg=C['bg_surface'])
        title_area.pack(side=tk.LEFT, padx=24, pady=12)

        # Title row with app icon
        title_row = tk.Frame(title_area, bg=C['bg_surface'])
        title_row.pack(anchor=tk.W)
        if 'header' in self._icon_refs:
            tk.Label(title_row,
                image=self._icon_refs['header'],
                bg=C['bg_surface'],
            ).pack(side=tk.LEFT, padx=(0, 8))
        else:
            tk.Label(title_row,
                text="\u25c6",
                font=(ff, 18),
                fg=C['accent'],
                bg=C['bg_surface'],
            ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(title_row,
            text="Bulk PDF Generator",
            font=(ff, 22, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_surface'],
        ).pack(side=tk.LEFT)

        tk.Label(title_area,
            text="Generate filled PDFs from spreadsheet data",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
        ).pack(anchor=tk.W, padx=(30, 0))

        info_area = tk.Frame(header, bg=C['bg_surface'])
        info_area.pack(side=tk.RIGHT, padx=24, pady=12)

        # School name (clickable to edit)
        school_text = ""
        if self.settings.school_configured:
            school_text = f"{self.settings.school_name}  |  {self.settings.school_year}"
        self.header_school = tk.Label(info_area,
            text=school_text,
            font=(ff, 10, 'bold'),
            fg=C['accent'],
            bg=C['bg_surface'],
            cursor="hand2",
        )
        self.header_school.pack(anchor=tk.E)
        self.header_school.bind("<Button-1>", lambda e: self.prompt_school_setup())

        self.header_status = tk.Label(info_area,
            text="No template loaded",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
        )
        self.header_status.pack(anchor=tk.E)

        # Accent stripe divider
        tk.Frame(main_frame, bg=C['accent'], height=3).pack(fill=tk.X)

        # ── Content area ──
        content_frame = ttk.Frame(main_frame, padding=str(SPACING['page_padding']))
        content_frame.pack(fill=tk.BOTH, expand=True)

        # Notebook (tabs)
        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Tab 0 uses a plain frame (Text widget has its own scrollbar)
        self.tab0_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab0_frame, text="  Getting Started  ")

        # Tabs 1-3 use ScrollableFrame
        self.tab1_container = ScrollableFrame(self.notebook)
        self.tab2_container = ScrollableFrame(self.notebook)
        self.tab3_container = ScrollableFrame(self.notebook)

        self.notebook.add(self.tab1_container, text="  1  Analyze Template  ")
        self.notebook.add(self.tab2_container, text="  2  Map Fields  ")
        self.notebook.add(self.tab3_container, text="  3  Generate PDFs  ")

        # About tab (plain frame, right-hand side)
        self.tab_about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_about_frame, text="  About  ")

        # Shorthand for actual UI frames
        self.tab1 = self.tab1_container.scrollable_frame
        self.tab2 = self.tab2_container.scrollable_frame
        self.tab3 = self.tab3_container.scrollable_frame

        # ── Status Bar ──
        tk.Frame(main_frame, bg=C['border_subtle'], height=1).pack(fill=tk.X, side=tk.BOTTOM)
        status_frame = tk.Frame(main_frame, bg=C['bg_surface'], height=32)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        status_frame.pack_propagate(False)

        self.status_label = tk.Label(status_frame,
            text="Ready",
            font=(ff, 9),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
            anchor=tk.W,
        )
        self.status_label.pack(side=tk.LEFT, padx=12)

        self.status_template = tk.Label(status_frame,
            text="No template loaded",
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_surface'],
            anchor=tk.E,
        )
        self.status_template.pack(side=tk.RIGHT, padx=12)

        # Setup each tab
        self.setup_tab0_getting_started()
        self.setup_tab1_analyze()
        self.setup_tab2_mapping()
        self.setup_tab3_generate()
        self.setup_tab_about()

        # Disable Tab 2 (placeholder until Phase 3)
        self.notebook.tab(2, state='disabled')

    # ========== UI HELPERS ==========

    def create_section(self, parent, title, subtitle=None, expand=False):
        """Create a card-style section with title label and bordered content area.

        Returns the inner content frame to pack widgets into.
        """
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Section wrapper
        wrapper = ttk.Frame(parent)
        fill_mode = tk.BOTH if expand else tk.X
        wrapper.pack(fill=fill_mode, expand=expand, pady=(0, SPACING['section_gap']))

        # Title label
        tk.Label(wrapper,
            text=title,
            font=(ff, 13, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_base'],
            anchor=tk.W,
        ).pack(anchor=tk.W, pady=(0, 6))

        # Optional subtitle
        if subtitle:
            tk.Label(wrapper,
                text=subtitle,
                font=(ff, 10),
                fg=C['text_secondary'],
                bg=C['bg_base'],
                anchor=tk.W,
            ).pack(anchor=tk.W, pady=(0, 6))

        # Border frame (simulates 1px border via bg color + padding)
        border_frame = tk.Frame(wrapper, bg=C['border_subtle'], padx=1, pady=1)
        border_frame.pack(fill=fill_mode, expand=expand)

        # Inner content frame
        inner = tk.Frame(border_frame, bg=C['bg_surface'],
                         padx=SPACING['inner_padding'], pady=SPACING['inner_padding'])
        inner.pack(fill=fill_mode, expand=expand)

        return inner

    # ========== TAB 0: GETTING STARTED ==========

    def setup_tab0_getting_started(self):
        """Build the Getting Started tab with rendered markdown content."""
        C = COLORS

        # Text widget + scrollbar
        text_frame = ttk.Frame(self.tab0_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            bg=C['bg_base'],
            fg=C['text_primary'],
            insertbackground=C['text_primary'],
            selectbackground=C['accent_subtle'],
            selectforeground=C['text_primary'],
            borderwidth=0,
            highlightthickness=0,
            padx=20,
            pady=15,
            yscrollcommand=scrollbar.set,
            cursor='arrow',
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)

        # Load and render the markdown file
        md_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'getting_started.md')
        try:
            load_and_render(text_widget, md_path)
        except FileNotFoundError:
            text_widget.insert(tk.END, "Getting Started content not found.\n\n"
                               "Expected file: getting_started.md\n"
                               "Place it in the same folder as the application.")
            text_widget.config(state=tk.DISABLED)
        except (PermissionError, UnicodeDecodeError, OSError) as e:
            text_widget.insert(tk.END, f"Could not load Getting Started content:\n{e}")
            text_widget.config(state=tk.DISABLED)

    def setup_tab_about(self):
        """Build the About tab with developer info and project links."""
        import webbrowser
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Outer wrapper centres content vertically and horizontally
        outer = tk.Frame(self.tab_about_frame, bg=C['bg_base'])
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        # Centred card
        card_border = tk.Frame(outer, bg=C['border_subtle'], padx=1, pady=1)
        card_border.grid(row=0, column=0)  # centres in the expanded cell
        card = tk.Frame(card_border, bg=C['bg_surface'], padx=48, pady=40)
        card.pack()

        # App icon + title
        if 'about' in self._icon_refs:
            tk.Label(card, image=self._icon_refs['about'],
                     bg=C['bg_surface']).pack(pady=(0, 8))
        else:
            tk.Label(card, text="\u25c6", font=(ff, 36), fg=C['accent'],
                     bg=C['bg_surface']).pack(pady=(0, 4))
        tk.Label(card, text="Bulk PDF Generator", font=(ff, 22, 'bold'),
                 fg=C['text_primary'], bg=C['bg_surface']).pack()
        tk.Label(card, text="Generate filled PDFs from spreadsheet data",
                 font=(ff, 11), fg=C['text_secondary'],
                 bg=C['bg_surface']).pack(pady=(2, 20))

        # Divider
        tk.Frame(card, bg=C['border_subtle'], height=1).pack(fill=tk.X, pady=(0, 20))

        # Mission statement
        mission = (
            "Designed to take the pain out of complicated, repetitive\n"
            "PDF form-filling tasks in schools \u2014 turning hours of manual\n"
            "data entry into a single click, so staff can focus on the\n"
            "work that actually matters."
        )
        tk.Label(card, text=mission, font=(ff, 11), fg=C['text_primary'],
                 bg=C['bg_surface'], justify=tk.CENTER,
                 wraplength=420).pack(pady=(0, 24))

        # Developer info
        tk.Label(card, text="Developed by", font=(ff, 10),
                 fg=C['text_tertiary'], bg=C['bg_surface']).pack()
        tk.Label(card, text="Dave Armstrong", font=(ff, 14, 'bold'),
                 fg=C['text_primary'], bg=C['bg_surface']).pack(pady=(2, 2))
        tk.Label(card, text="Victorian Department of Education", font=(ff, 10),
                 fg=C['text_secondary'], bg=C['bg_surface']).pack(pady=(0, 16))

        # Email link
        email_label = tk.Label(card, text="Dave.Armstrong@education.vic.gov.au",
                               font=(ff, 11, 'underline'), fg=C['info'],
                               bg=C['bg_surface'], cursor='hand2')
        email_label.pack(pady=(0, 6))
        email_label.bind('<Button-1>',
                         lambda e: webbrowser.open('mailto:Dave.Armstrong@education.vic.gov.au'))

        # GitHub link
        github_label = tk.Label(card, text="github.com/mrdavearms/VCAA-PDF-Generator",
                                font=(ff, 11, 'underline'), fg=C['info'],
                                bg=C['bg_surface'], cursor='hand2')
        github_label.pack(pady=(0, 20))
        github_label.bind('<Button-1>',
                          lambda e: webbrowser.open('https://github.com/mrdavearms/VCAA-PDF-Generator'))

        # Version
        tk.Label(card, text="v2.0", font=(ff, 10),
                 fg=C['text_tertiary'], bg=C['bg_surface']).pack()

    def update_status(self, message, level='info'):
        """Update the status bar message."""
        color_map = {
            'info': COLORS['text_secondary'],
            'success': COLORS['success'],
            'warning': COLORS['warning'],
            'error': COLORS['error'],
        }
        self.status_label.config(
            text=message,
            fg=color_map.get(level, COLORS['text_secondary'])
        )

    # ========== TAB 1: ANALYZE TEMPLATE ==========

    def setup_tab1_analyze(self):
        """Setup Tab 1: PDF Template Analysis."""
        tab = self.tab1

        # Container with padding
        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        # Section: Load Template
        load_inner = self.create_section(container, "Load Template")

        # PDF selection row
        pdf_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        pdf_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(pdf_row, text="PDF Template:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(pdf_row, textvariable=self.pdf_template_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(pdf_row, text="Browse...", command=self.select_pdf_tab1, width=10).pack(side=tk.LEFT)

        # Template name row
        name_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        name_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(name_row, text="Template Name:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        self.template_name_var = tk.StringVar()
        ttk.Entry(name_row, textvariable=self.template_name_var, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)

        # Naming options
        naming_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        naming_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']), padx=(144, 0))
        self.naming_option_var = tk.StringVar(value="auto")
        ttk.Radiobutton(naming_row, text="Auto-name from PDF", variable=self.naming_option_var, value="auto", style='Surface.TRadiobutton').pack(side=tk.LEFT, padx=(0, 12))
        ttk.Radiobutton(naming_row, text="Custom name", variable=self.naming_option_var, value="custom", style='Surface.TRadiobutton').pack(side=tk.LEFT)

        # Recent templates dropdown
        recent_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        recent_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(recent_row, text="Recent Templates:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        self.recent_templates_var = tk.StringVar()
        self.recent_templates_combo = ttk.Combobox(recent_row, textvariable=self.recent_templates_var, state="readonly", width=47)
        self.recent_templates_combo.pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(recent_row, text="Load", command=self.load_recent_template, width=10).pack(side=tk.LEFT)

        self.populate_recent_templates()

        # Analyze button
        btn_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        btn_row.pack(fill=tk.X, pady=(SPACING['element_gap'], 0))
        ttk.Button(btn_row, text="Analyze Fields", command=self.analyze_pdf_fields, style='Accent.TButton').pack()

        # Analysis results section
        results_inner = self.create_section(container, "Analysis Results", expand=True)

        # Stats row
        self.stats_label = tk.Label(results_inner, text="No analysis performed yet",
                                    font=font(10), fg=COLORS['text_secondary'], bg=COLORS['bg_surface'])
        self.stats_label.pack(pady=(0, 8))

        # Fields table
        table_frame = tk.Frame(results_inner, bg=COLORS['bg_surface'])
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('field_name', 'type', 'page', 'length')
        self.fields_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        self.fields_tree.heading('field_name', text='Field Name')
        self.fields_tree.heading('type', text='Type')
        self.fields_tree.heading('page', text='Page')
        self.fields_tree.heading('length', text='Length/Size')

        self.fields_tree.column('field_name', width=250)
        self.fields_tree.column('type', width=120)
        self.fields_tree.column('page', width=60)
        self.fields_tree.column('length', width=100)

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.fields_tree.yview)
        self.fields_tree.configure(yscrollcommand=scrollbar.set)

        self.fields_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Setup treeview tags and hover
        setup_treeview_tags(self.fields_tree)
        bind_treeview_hover(self.fields_tree)

        # Bind selection event for field preview
        self.fields_tree.bind('<<TreeviewSelect>>', self.on_field_selected)

        # Visual Preview section (nested inside results card)
        preview_header = tk.Frame(results_inner, bg=COLORS['bg_surface'])
        preview_header.pack(fill=tk.X, pady=(SPACING['element_gap'], 6))

        tk.Label(preview_header, text="Field Preview",
                 font=font(11, 'bold'), fg=COLORS['text_primary'],
                 bg=COLORS['bg_surface'], anchor=tk.W).pack(side=tk.LEFT)

        # Zoom controls
        zoom_frame = tk.Frame(preview_header, bg=COLORS['bg_surface'])
        zoom_frame.pack(side=tk.RIGHT)
        self.zoom_level = 1.0
        self._preview_raw_img = None  # Store unscaled PIL image for zoom

        ttk.Button(zoom_frame, text="\u2212", width=3,
                   command=lambda: self._zoom_preview(-0.25)).pack(side=tk.LEFT, padx=2)
        self.zoom_label = tk.Label(zoom_frame, text="100%", width=5,
                                   font=font(10), fg=COLORS['text_secondary'],
                                   bg=COLORS['bg_surface'], anchor=tk.CENTER)
        self.zoom_label.pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="+", width=3,
                   command=lambda: self._zoom_preview(0.25)).pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="Fit", width=4,
                   command=lambda: self._zoom_preview(0, fit=True)).pack(side=tk.LEFT, padx=(6, 0))

        tk.Label(results_inner, text="Click a field above to preview \u2022 Use +/\u2212 to zoom \u2022 Scroll to pan",
                 font=font(9), fg=COLORS['text_secondary'], bg=COLORS['bg_surface'], anchor=tk.W).pack(fill=tk.X, pady=(0, 6))

        # Preview canvas (scrollable for zoom)
        canvas_frame = tk.Frame(results_inner, bg=COLORS['canvas_border'])
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.preview_canvas = tk.Canvas(
            canvas_frame, width=600, height=300,
            bg=COLORS['canvas_bg'],
            highlightthickness=0,
            xscrollincrement=1, yscrollincrement=1,
        )
        self.preview_canvas.pack(fill=tk.BOTH, expand=True)

        # Mouse wheel scrolling on canvas for panning when zoomed
        def _on_canvas_scroll(event):
            if self.zoom_level > 1.0:
                self.preview_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.preview_canvas.bind('<MouseWheel>', _on_canvas_scroll)

        self.preview_image = None
        self.preview_generator = None

        # Action buttons row
        action_frame = tk.Frame(container, bg=COLORS['bg_base'])
        action_frame.pack(fill=tk.X, pady=(0, SPACING['element_gap']))

        ttk.Button(action_frame, text="Export Mapping File (.xlsx)", command=self.export_mapping_file).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(action_frame, text="Save Template Config", command=self.save_template_config, style='Accent.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(action_frame, text="Skip to Generate PDFs \u2192", command=lambda: self.notebook.select(3)).pack(side=tk.RIGHT)

    def select_pdf_tab1(self):
        """Select PDF file in Tab 1."""
        filepath = filedialog.askopenfilename(
            title="Select PDF Template",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filepath:
            self.pdf_template_path.set(filepath)
            # Auto-generate template name
            if self.naming_option_var.get() == "auto":
                suggested_name = auto_name_template(os.path.basename(filepath))
                self.template_name_var.set(suggested_name)

    def populate_recent_templates(self):
        """Populate recent templates dropdown."""
        templates_dir = Path(self.settings.templates_directory)
        if not templates_dir.exists():
            return

        template_files = list(templates_dir.glob("*.json"))
        template_names = [f.stem for f in template_files]

        self.recent_templates_combo['values'] = template_names

    def load_recent_template(self):
        """Load a template from the recent list."""
        template_name = self.recent_templates_var.get()
        if not template_name:
            return

        template_path = os.path.join(self.settings.templates_directory, f"{template_name}.json")
        if os.path.exists(template_path):
            self.load_template_config(template_path)

    def load_template_from_file(self):
        """Load template config from file dialog."""
        filepath = filedialog.askopenfilename(
            title="Select Template Configuration",
            initialdir=self.settings.templates_directory,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filepath:
            self.load_template_config(filepath)

    def load_template_config(self, filepath: str):
        """Load template configuration."""
        try:
            self.current_template = TemplateConfig.from_file(filepath)
            self.pdf_template_path.set(self.current_template.pdf_path)
            self.template_name_var.set(self.current_template.template_name)

            self.update_status(f"Template loaded: {self.current_template.template_name}", 'success')
            self.header_status.config(text=f"Template: {self.current_template.template_name}")
            self.status_template.config(text=self.current_template.template_name)
            messagebox.showinfo("Template Loaded", f"Loaded template: {self.current_template.template_name}")

            # Switch to Tab 2 or 3
            self.notebook.select(3)  # Go to Generate tab

        except Exception as e:
            self.update_status(f"Failed to load template", 'error')
            messagebox.showerror("Error", f"Failed to load template:\n{str(e)}")

    def analyze_pdf_fields(self):
        """Analyze PDF fields and detect combed fields."""
        pdf_path = self.pdf_template_path.get()

        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "Please select a valid PDF template file.")
            return

        try:
            # Get template name (show dialog if custom)
            if self.naming_option_var.get() == "custom":
                suggested_name = self.template_name_var.get() or auto_name_template(os.path.basename(pdf_path))
                dialog = TemplateNameDialog(self.root, suggested_name)
                self.root.wait_window(dialog)

                if dialog.result is None:
                    return  # User cancelled

                template_name = dialog.result
            else:
                template_name = auto_name_template(os.path.basename(pdf_path))

            self.template_name_var.set(template_name)

            # Analyze PDF
            with PDFAnalyzer(pdf_path) as analyzer:
                self.analyzed_fields = analyzer.analyze_fields()
                field_stats = analyzer.get_field_statistics(self.analyzed_fields)

            # Close any existing preview generator before opening a new one
            self._close_preview_generator()

            # Initialize preview generator
            cache_dir = os.path.join(self.settings.templates_directory, '.preview_cache')
            self.preview_generator = VisualPreviewGenerator(pdf_path, cache_dir)
            self.preview_generator.__enter__()  # Open the PDF

            # Update UI
            self.display_analyzed_fields()

            # Update stats label
            total = len(self.analyzed_fields)
            combed = sum(1 for f in self.analyzed_fields if f.is_combed)
            pages = len(set(f.page for f in self.analyzed_fields))

            self.stats_label.config(
                text=f"Template: {template_name}  |  Total Fields: {total}  |  Pages: {pages}  |  Combed: {combed}"
            )

            # Update status bar and header
            self.update_status(f"Analysis complete: {total} fields found ({combed} combed)", 'success')
            self.header_status.config(text=f"Template: {template_name}")
            self.status_template.config(text=f"{template_name} ({total} fields)")

            messagebox.showinfo("Analysis Complete", f"Found {total} fields ({combed} combed fields)")

        except Exception as e:
            self._close_preview_generator()  # Guarantee cleanup on error
            self.update_status(f"Analysis failed: {str(e)}", 'error')
            messagebox.showerror("Error", f"Failed to analyze PDF:\n{str(e)}")

    def display_analyzed_fields(self):
        """Display analyzed fields in the table."""
        # Clear existing
        for item in self.fields_tree.get_children():
            self.fields_tree.delete(item)

        # Add fields with alternating row colors
        for i, field in enumerate(self.analyzed_fields):
            length_display = f"{field.length} chars" if field.is_combed else "Single"
            row_tag = 'odd' if i % 2 else 'even'
            tags = (row_tag, 'combed') if field.is_combed else (row_tag,)

            self.fields_tree.insert('', tk.END, values=(
                field.field_name,
                field.field_type,
                field.page,
                length_display
            ), tags=tags)

    def _zoom_preview(self, delta, fit=False):
        """Adjust preview zoom level and re-render."""
        if fit:
            self.zoom_level = 1.0
        else:
            self.zoom_level = max(0.5, min(4.0, self.zoom_level + delta))

        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")
        self._render_preview_at_zoom()

    def _render_preview_at_zoom(self):
        """Re-render the stored raw preview image at the current zoom level."""
        if self._preview_raw_img is None:
            return

        try:
            canvas_width = self.preview_canvas.winfo_width()
            canvas_height = self.preview_canvas.winfo_height()
            if canvas_width <= 1:
                canvas_width = 600
            if canvas_height <= 1:
                canvas_height = 300

            img_width, img_height = self._preview_raw_img.size

            # Base scale: fit to canvas
            base_scale = min(canvas_width / img_width, canvas_height / img_height) * 0.95
            scale = base_scale * self.zoom_level

            new_width = max(1, int(img_width * scale))
            new_height = max(1, int(img_height * scale))

            resized = self._preview_raw_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.preview_image = ImageTk.PhotoImage(resized)

            self.preview_canvas.delete("all")

            if self.zoom_level > 1.0:
                # Enable scrolling for zoomed images
                self.preview_canvas.config(scrollregion=(0, 0, new_width, new_height))
                self.preview_canvas.create_image(new_width // 2, new_height // 2,
                                                 image=self.preview_image, anchor=tk.CENTER)
            else:
                # Reset scroll region to canvas size when fitting
                self.preview_canvas.config(scrollregion=(0, 0, canvas_width, canvas_height))
                self.preview_canvas.create_image(canvas_width // 2, canvas_height // 2,
                                                 image=self.preview_image, anchor=tk.CENTER)
        except Exception as e:
            print(f"Zoom render error: {e}")

    def on_field_selected(self, event):
        """Handle field selection in Tab 1 - show visual preview."""
        if (not self.analyzed_fields
                or not self.preview_generator
                or not self.preview_generator.doc):
            return

        selection = self.fields_tree.selection()
        if not selection:
            return

        # Get selected item index
        item = selection[0]
        item_index = self.fields_tree.index(item)

        if item_index >= len(self.analyzed_fields):
            return

        # Get the selected field
        field = self.analyzed_fields[item_index]

        try:
            # Generate preview at higher DPI for zoom quality
            preview_img = self.preview_generator.generate_field_preview(field, dpi=200)

            # Store raw image for zoom
            self._preview_raw_img = preview_img
            self.zoom_level = 1.0
            self.zoom_label.config(text="100%")

            # Render at current zoom (handles canvas drawing)
            self._render_preview_at_zoom()

        except Exception as e:
            print(f"Preview error: {e}")
            self.preview_canvas.delete("all")
            self.preview_canvas.create_text(
                300, 150,
                text=f"Preview unavailable\n{str(e)}",
                fill=COLORS['error'],
                font=font(12),
            )

    def export_mapping_file(self):
        """Export field mapping to Excel file with formatted worksheets.

        Creates three sheets:
        - Field Mapping: detailed reference of all PDF fields
        - Data Entry: ready-to-use data entry sheet with field names as column headers
        - Instructions: how-to guide for using the file
        All cells have wrap-text formatting applied.
        """
        if not self.analyzed_fields:
            messagebox.showwarning("No Data", "Please analyze a PDF template first.")
            return

        template_name = self.template_name_var.get() or "template"
        suggested_filename = f"{template_name}_mapping.xlsx"

        filepath = filedialog.asksaveasfilename(
            title="Save Mapping File",
            initialdir=self.settings.templates_directory,
            initialfile=suggested_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not filepath:
            return

        try:
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            # Build field data
            data = []
            for field in self.analyzed_fields:
                excel_col = self.smart_guess_excel_column(field.field_name)
                data.append({
                    'PDF_Field_Name': field.field_name,
                    'Excel_Column_Name': excel_col,
                    'Field_Type': field.field_type,
                    'Page': field.page,
                    'Required': 'Yes' if field.field_name.lower() in [
                        'surname', 'first_name', 'first name',
                        'vcaa_number', 'vcaa student number'
                    ] else 'No',
                    'Length': f"{field.length} chars" if field.is_combed else '-',
                    'Notes': self.generate_field_notes(field),
                })

            df = pd.DataFrame(data)

            # Data Entry columns: use the Excel column names as headers
            data_entry_cols = [d['Excel_Column_Name'] for d in data if d['Excel_Column_Name']]
            df_data_entry = pd.DataFrame(columns=data_entry_cols)

            # Write sheets via pandas, then format with openpyxl
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Field Mapping', index=False)
                df_data_entry.to_excel(writer, sheet_name='Data Entry', index=False)

                # Instructions sheet
                instructions = pd.DataFrame({
                    'Bulk PDF Generator - Field Mapping Guide': [
                        '',
                        'HOW TO USE THIS FILE:',
                        '1. Review the "Field Mapping" sheet',
                        '2. Update any "Excel_Column_Name" values to match your preferred headers',
                        '3. The "Data Entry" sheet is where you can start typing student data',
                        '4. Save this file',
                        '5. Return to the app and load this file in Tab 3 (Generate PDFs)',
                        '',
                        'IMPORTANT NOTES:',
                        '- The "Data Entry" sheet headers are automatically generated',
                        '- If you change column names in "Field Mapping", you should also update them in "Data Entry"',
                        '- Combed fields will auto-split text (e.g., "John" → J-o-h-n)',
                        '- The app uses the "Data Entry" sheet for PDF generation',
                        '',
                        'For help: See the Getting Started tab in the app',
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False, header=False)

                wb = writer.book

                # -- Shared styles --
                wrap_align = Alignment(wrap_text=True, vertical='top')
                header_font = Font(bold=True, size=11)
                header_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                thin_border = Border(
                    bottom=Side(style='thin', color='E0E0E3'),
                )

                # ── Format: Field Mapping sheet ──
                ws_map = wb['Field Mapping']
                for row in ws_map.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_align
                # Bold + coloured header row
                for cell in ws_map[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                # Auto-width columns (min 14, max 40)
                for col in ws_map.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    adjusted = min(40, max(14, max_len + 4))
                    ws_map.column_dimensions[col[0].column_letter].width = adjusted

                # ── Format: Data Entry sheet ──
                ws_entry = wb['Data Entry']
                # Apply wrap text to ALL cells (headers + 50 empty rows for data entry)
                for r in range(1, 52):
                    for c in range(1, len(data_entry_cols) + 1):
                        cell = ws_entry.cell(row=r, column=c)
                        cell.alignment = wrap_align
                # Bold + coloured header row
                for cell in ws_entry[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                # Set column widths (min 16 based on header text)
                for idx, col_name in enumerate(data_entry_cols, 1):
                    width = min(30, max(16, len(col_name) + 4))
                    ws_entry.column_dimensions[ws_entry.cell(row=1, column=idx).column_letter].width = width
                # Freeze header row so it stays visible during scrolling
                ws_entry.freeze_panes = 'A2'

                # ── Format: Instructions sheet ──
                ws_instr = wb['Instructions']
                for row in ws_instr.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_align
                ws_instr.column_dimensions['A'].width = 80

            messagebox.showinfo("Export Successful", f"Mapping file saved to:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export mapping file:\n{str(e)}")

    def smart_guess_excel_column(self, pdf_field_name: str) -> str:
        """Generate a smart guess for Excel column name."""
        # Convert underscores to spaces
        name = pdf_field_name.replace('_', ' ')

        # Known mappings
        mappings = {
            'first name': 'First name',
            'surname': 'surname',
            'vcaa number': 'VCAA student number',
            'vcaa student number': 'VCAA student number',
            'school name': 'School name',
            'vcaa school code': 'VCAA school code',
        }

        lower_name = name.lower()
        if lower_name in mappings:
            return mappings[lower_name]

        return name

    def generate_field_notes(self, field: PDFField) -> str:
        """Generate helpful notes for a field."""
        notes = []

        if field.field_name.lower() in ['surname', 'first_name', 'first name']:
            notes.append('Used for filename')

        if field.field_name.lower() in ['surname', 'first name', 'vcaa student number', 'vcaa_number']:
            notes.append('Critical field')

        if field.is_combed and 'dob' in field.field_name.lower():
            notes.append('Auto-format date')

        return ', '.join(notes) if notes else ''

    def save_template_config(self):
        """Save template configuration to JSON."""
        if not self.analyzed_fields:
            messagebox.showwarning("No Data", "Please analyze a PDF template first.")
            return

        template_name = self.template_name_var.get()
        if not template_name:
            messagebox.showwarning("No Name", "Please provide a template name.")
            return

        try:
            # Create config
            field_stats = {}
            for field in self.analyzed_fields:
                key = field.field_type.lower().replace('-', '_').replace(' ', '_')
                field_stats[key] = field_stats.get(key, 0) + 1

            config = TemplateConfig(
                template_name=template_name,
                pdf_filename=os.path.basename(self.pdf_template_path.get()),
                pdf_path=self.pdf_template_path.get(),
                created_date=datetime.now().isoformat(),
                last_used=datetime.now().isoformat(),
                total_fields=len(self.analyzed_fields),
                field_types=field_stats,
                mapping_file=f"{template_name}_mapping.xlsx",
                use_auto_matching=True,
                critical_fields=['surname', 'First name', 'VCAA student number'],
                notes="",
                version="2.0"
            )

            # Save to file
            config_path = os.path.join(self.settings.templates_directory, f"{template_name}.json")
            config.save_to_file(config_path)

            self.current_template = config

            messagebox.showinfo("Saved", f"Template configuration saved to:\n{config_path}")

            # Refresh recent templates
            self.populate_recent_templates()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template config:\n{str(e)}")

    # ========== TAB 2: MAP FIELDS ==========

    def setup_tab2_mapping(self):
        """Setup Tab 2: Field Mapping (Placeholder for now)."""
        tab = self.tab2

        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            container,
            text="Field Mapping",
            style='Title.TLabel',
        ).pack(pady=(20, 10))

        ttk.Label(
            container,
            text="This tab will allow manual field mapping configuration.",
            style='Subtitle.TLabel',
        ).pack(pady=SPACING['element_gap'])

        ttk.Label(
            container,
            text="Auto-matching is enabled by default",
            style='Success.TLabel',
        ).pack(pady=5)

        ttk.Label(
            container,
            text="Coming in Phase 3...",
            style='Muted.TLabel',
        ).pack(pady=20)

    # ========== TAB 3: GENERATE PDFs (From original app) ==========

    def setup_tab3_generate(self):
        """Setup Tab 3: PDF Generation (original functionality)."""
        tab = self.tab3

        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        # Template selection bar
        template_frame = tk.Frame(container, bg=COLORS['bg_base'])
        template_frame.pack(fill=tk.X, pady=(0, SPACING['section_gap']))

        tk.Label(template_frame, text="Template:", font=font(10),
                 fg=COLORS['text_secondary'], bg=COLORS['bg_base']).pack(side=tk.LEFT, padx=(0, 10))

        self.selected_template_var = tk.StringVar()
        template_combo = ttk.Combobox(template_frame, textvariable=self.selected_template_var, state="readonly", width=40)
        template_combo.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(template_frame, text="Change Template", command=self.change_template_tab3).pack(side=tk.LEFT, padx=(0, 8))

        self.matching_status_label = ttk.Label(template_frame, text="Auto-matching enabled", style='Success.TLabel')
        self.matching_status_label.pack(side=tk.LEFT, padx=10)

        # File Selection section
        file_inner = self.create_section(container, "Select Files")

        # PDF Template selection
        pdf_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        pdf_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(pdf_row, text="PDF Template:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(pdf_row, textvariable=self.pdf_template_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(pdf_row, text="Browse...", command=self.select_pdf_tab3, width=10).pack(side=tk.LEFT)

        # Excel file selection
        excel_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        excel_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(excel_row, text="Excel Data File:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(excel_row, textvariable=self.excel_file_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(excel_row, text="Browse...", command=self.select_excel_tab3, width=10).pack(side=tk.LEFT)

        # Output folder selection
        output_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        output_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(output_row, text="Output Folder:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(output_row, textvariable=self.output_dir_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(output_row, text="Browse...", command=self.select_output_dir_tab3, width=10).pack(side=tk.LEFT)
        tk.Label(output_row, text="(optional)", font=font(9),
                 fg=COLORS['text_tertiary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT, padx=(6, 0))

        # Load button
        btn_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        btn_row.pack(fill=tk.X, pady=(SPACING['element_gap'], 0))
        ttk.Button(btn_row, text="Load & Preview Data", command=self.load_data_tab3, style='Accent.TButton').pack()

        # Validation section
        validation_inner = self.create_section(container, "Validation")

        self.validation_text_tab3 = tk.Text(
            validation_inner, height=3, wrap=tk.WORD, state=tk.DISABLED,
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['text_primary'],
            selectbackground=COLORS['accent_subtle'],
            selectforeground=COLORS['text_primary'],
            relief='flat',
            borderwidth=0,
            padx=10,
            pady=8,
            font=font(10),
        )
        self.validation_text_tab3.pack(fill=tk.X)
        # Keep a reference for pack_forget compatibility
        self.validation_frame_tab3 = validation_inner.master.master

        # Selection Controls
        selection_frame = tk.Frame(container, bg=COLORS['bg_base'])
        selection_frame.pack(fill=tk.X, pady=(0, SPACING['element_gap']))

        tk.Label(selection_frame, text="Select students to process:", font=font(11),
                 fg=COLORS['text_primary'], bg=COLORS['bg_base']).pack(side=tk.LEFT)

        ttk.Button(selection_frame, text="Select All", command=self.select_all_tab3).pack(side=tk.LEFT, padx=(15, 5))
        ttk.Button(selection_frame, text="Deselect All", command=self.deselect_all_tab3).pack(side=tk.LEFT, padx=(0, 5))

        self.selection_count_label_tab3 = ttk.Label(selection_frame, text="", style='Secondary.TLabel')
        self.selection_count_label_tab3.pack(side=tk.RIGHT)

        # Student Preview section
        preview_inner = self.create_section(container, "Student Preview",
                                            subtitle="Click rows to select or deselect", expand=True)

        # Treeview for student list
        columns = ('selected', 'row', 'surname', 'first_name', 'student_number', 'status')
        self.tree_tab3 = ttk.Treeview(preview_inner, columns=columns, show='headings', height=12)

        self.tree_tab3.heading('selected', text='Sel')
        self.tree_tab3.heading('row', text='#')
        self.tree_tab3.heading('surname', text='Surname')
        self.tree_tab3.heading('first_name', text='First Name')
        self.tree_tab3.heading('student_number', text='Student Number')
        self.tree_tab3.heading('status', text='Status')

        self.tree_tab3.column('selected', width=40, anchor='center')
        self.tree_tab3.column('row', width=40, anchor='center')
        self.tree_tab3.column('surname', width=140)
        self.tree_tab3.column('first_name', width=140)
        self.tree_tab3.column('student_number', width=110)
        self.tree_tab3.column('status', width=180)

        # Bind click event for toggling selection
        self.tree_tab3.bind('<ButtonRelease-1>', self.toggle_selection_tab3)

        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(preview_inner, orient=tk.VERTICAL, command=self.tree_tab3.yview)
        self.tree_tab3.configure(yscrollcommand=scrollbar.set)

        self.tree_tab3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Setup treeview tags and hover
        setup_treeview_tags(self.tree_tab3)
        bind_treeview_hover(self.tree_tab3)

        # Summary label
        self.summary_label_tab3 = ttk.Label(container, text="No data loaded", style='Secondary.TLabel')
        self.summary_label_tab3.pack(pady=(8, 4))

        # Progress Frame
        progress_frame = tk.Frame(container, bg=COLORS['bg_base'])
        progress_frame.pack(fill=tk.X, pady=(0, SPACING['section_gap']))

        self.progress_var_tab3 = tk.DoubleVar()
        self.progress_bar_tab3 = ttk.Progressbar(progress_frame, variable=self.progress_var_tab3, maximum=100)
        self.progress_bar_tab3.pack(fill=tk.X, pady=(4, 6))

        self.progress_label_tab3 = ttk.Label(progress_frame, text="", style='Secondary.TLabel')
        self.progress_label_tab3.pack()

        # Generate Button (large CTA)
        self.generate_btn_tab3 = ttk.Button(
            container,
            text="Generate PDFs for Selected Students",
            command=self.start_generation_tab3,
            state=tk.DISABLED,
            style='BigAccent.TButton',
        )
        self.generate_btn_tab3.pack(pady=SPACING['section_gap'])

    def select_pdf_tab3(self):
        """Select PDF in Tab 3."""
        filepath = filedialog.askopenfilename(
            title="Select PDF Template",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filepath:
            self.pdf_template_path.set(filepath)

    def select_excel_tab3(self):
        """Select Excel file in Tab 3."""
        filepath = filedialog.askopenfilename(
            title="Select Student Data Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            self.excel_file_path.set(filepath)

    def select_output_dir_tab3(self):
        """Select output directory for generated PDFs."""
        initial_dir = self.output_dir_path.get() or os.path.dirname(self.excel_file_path.get() or '')
        dirpath = filedialog.askdirectory(
            title="Select Output Folder for Generated PDFs",
            initialdir=initial_dir or None,
        )
        if dirpath:
            self.output_dir_path.set(dirpath)

    def change_template_tab3(self):
        """Change template in Tab 3."""
        self.load_template_from_file()

    def load_data_tab3(self):
        """Load Excel data for Tab 3 (original functionality)."""
        pdf_path = self.pdf_template_path.get()
        excel_path = self.excel_file_path.get()

        # Validate paths
        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "Please select a valid PDF template file.")
            return

        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Error", "Please select a valid Excel data file.")
            return

        try:
            # Load PDF and get field names
            reader = PdfReader(pdf_path)
            self.pdf_fields = []  # Always reset before re-reading
            fields = reader.get_fields()
            if fields:
                self.pdf_fields = list(fields.keys())
            else:
                messagebox.showerror("Error", "The PDF template has no fillable form fields.")
                return

            # Load Excel data (case-insensitive extension check)
            if excel_path.lower().endswith('.csv'):
                try:
                    self.df = pd.read_csv(excel_path, encoding='utf-8-sig')
                except UnicodeDecodeError:
                    self.df = pd.read_csv(excel_path, encoding='latin-1')
            else:
                self.df = pd.read_excel(excel_path)

            # Clean column names (strip whitespace, lowercase for matching)
            self.df.columns = [str(col).strip() for col in self.df.columns]

            # Create lowercase mapping for field matching
            self.column_mapping = {col.lower(): col for col in self.df.columns}
            self.field_mapping = {field.lower(): field for field in self.pdf_fields}

            # Clear previous selections
            self.selected_rows = {}

            # Validate and show preview
            self.validate_data_tab3()
            self.show_preview_tab3()

            # Select all by default
            self.select_all_tab3()

            # Enable generate button
            self.generate_btn_tab3.config(state=tk.NORMAL)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data:\n{str(e)}")

    def validate_data_tab3(self):
        """Validate data for Tab 3."""
        warnings = []

        # Check each row for critical fields
        for idx, row in self.df.iterrows():
            row_warnings = []
            row_dict = {str(col).lower(): val for col, val in row.items()}

            for field in self.critical_fields:
                val = row_dict.get(field, '')
                if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                    row_warnings.append(field)

            if row_warnings:
                surname = row_dict.get('surname', f'Row {idx+1}')
                if pd.isna(surname) or str(surname).strip() == '':
                    surname = f'Row {idx+1}'
                warnings.append(f"• {surname}: Missing {', '.join(row_warnings)}")

        # Display warnings
        self.validation_text_tab3.config(state=tk.NORMAL)
        self.validation_text_tab3.delete(1.0, tk.END)

        if warnings:
            warning_text = f"{len(warnings)} student(s) have missing critical fields:\n"
            warning_text += "\n".join(warnings[:10])
            if len(warnings) > 10:
                warning_text += f"\n... and {len(warnings) - 10} more"
            self.validation_text_tab3.insert(1.0, warning_text)
            self.validation_text_tab3.config(fg=COLORS['warning'])
        else:
            self.validation_text_tab3.insert(1.0, "All students have required fields populated.")
            self.validation_text_tab3.config(fg=COLORS['success'])

        self.validation_text_tab3.config(state=tk.DISABLED)

    def show_preview_tab3(self):
        """Show preview for Tab 3."""
        # Clear existing items
        for item in self.tree_tab3.get_children():
            self.tree_tab3.delete(item)

        # Reset selected rows tracking
        self.selected_rows = {}

        # Add students to preview
        valid_count = 0
        warning_count = 0

        for idx, row in self.df.iterrows():
            row_dict = {str(col).lower(): val for col, val in row.items()}

            surname = str(row_dict.get('surname', '')).strip()
            first_name = str(row_dict.get('first name', '')).strip()
            student_num = str(row_dict.get('vcaa student number', '')).strip()

            # Skip completely empty rows
            if (pd.isna(row_dict.get('surname')) or surname == '' or surname.lower() == 'nan'):
                continue

            # Check status
            missing = []
            for field in self.critical_fields:
                val = row_dict.get(field, '')
                if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                    missing.append(field)

            if missing:
                status = f"Missing: {', '.join(missing)}"
                warning_count += 1
            else:
                status = "Ready"
                valid_count += 1

            # Clean display values
            if surname.lower() == 'nan':
                surname = ''
            if first_name.lower() == 'nan':
                first_name = ''
            if student_num.lower() == 'nan':
                student_num = ''

            # Build tags: alternating rows + status
            row_num = valid_count + warning_count
            row_tag = 'odd' if row_num % 2 else 'even'
            status_tag = 'warning_row' if missing else 'success_row'

            # Insert with empty checkbox initially
            item_id = self.tree_tab3.insert('', tk.END,
                values=('', idx+1, surname, first_name, student_num, status),
                tags=(row_tag, status_tag),
            )
            # Store mapping of item_id to dataframe index
            self.selected_rows[item_id] = {'index': idx, 'selected': False}

        # Update summary
        total = valid_count + warning_count
        self.summary_label_tab3.config(
            text=f"Loaded {total} students: {valid_count} ready, {warning_count} with warnings"
        )

        self.update_selection_count_tab3()

    def toggle_selection_tab3(self, event):
        """Toggle selection in Tab 3."""
        item = self.tree_tab3.identify_row(event.y)
        if item and item in self.selected_rows:
            # Toggle the selection
            self.selected_rows[item]['selected'] = not self.selected_rows[item]['selected']

            # Update the checkbox display
            current_values = list(self.tree_tab3.item(item, 'values'))
            current_values[0] = 'Yes' if self.selected_rows[item]['selected'] else ''
            self.tree_tab3.item(item, values=current_values)

            self.update_selection_count_tab3()

    def select_all_tab3(self):
        """Select all in Tab 3."""
        for item_id in self.selected_rows:
            self.selected_rows[item_id]['selected'] = True
            current_values = list(self.tree_tab3.item(item_id, 'values'))
            current_values[0] = 'Yes'
            self.tree_tab3.item(item_id, values=current_values)

        self.update_selection_count_tab3()

    def deselect_all_tab3(self):
        """Deselect all in Tab 3."""
        for item_id in self.selected_rows:
            self.selected_rows[item_id]['selected'] = False
            current_values = list(self.tree_tab3.item(item_id, 'values'))
            current_values[0] = ''
            self.tree_tab3.item(item_id, values=current_values)

        self.update_selection_count_tab3()

    def update_selection_count_tab3(self):
        """Update selection count for Tab 3."""
        selected = sum(1 for item in self.selected_rows.values() if item['selected'])
        total = len(self.selected_rows)
        self.selection_count_label_tab3.config(text=f"Selected: {selected} of {total}")

        # Update button text
        if selected == 0:
            self.generate_btn_tab3.config(text="Generate PDFs (none selected)", state=tk.DISABLED)
        elif selected == 1:
            self.generate_btn_tab3.config(text="Generate PDF for 1 Student", state=tk.NORMAL)
        else:
            self.generate_btn_tab3.config(text=f"Generate PDFs for {selected} Students", state=tk.NORMAL)

    def start_generation_tab3(self):
        """Start PDF generation for Tab 3.

        All shared state is snapshot here (main thread) so the worker
        thread never touches tkinter StringVars or mutable instance data.
        """
        # Check if any students are selected
        selected_count = sum(1 for item in self.selected_rows.values() if item['selected'])
        if selected_count == 0:
            messagebox.showwarning("No Selection", "Please select at least one student to generate PDFs.")
            return

        # ── Snapshot all shared state in the main thread ──
        ctx = {
            'pdf_path': self.pdf_template_path.get(),
            'excel_path': self.excel_file_path.get(),
            'df': self.df.copy(),
            'selected_indices': [
                item['index'] for item in self.selected_rows.values()
                if item['selected']
            ],
            'analyzed_fields': list(self.analyzed_fields),
            'pdf_fields': list(self.pdf_fields) if hasattr(self, 'pdf_fields') else [],
            'school_name': self.settings.school_name or "School",
            'school_year': self.settings.school_year or str(datetime.now().year),
            'combed_padding': self.settings.combed_field_padding,
            'combed_align': self.settings.combed_field_align,
            'output_dir': self.output_dir_path.get().strip() or '',
        }

        # Reset progress bar and disable button
        self.progress_var_tab3.set(0)
        self.generate_btn_tab3.config(state=tk.DISABLED)
        self.update_status(f"Generating PDFs for {selected_count} students...", 'info')

        # Start generation in background thread with snapshot
        thread = threading.Thread(target=self.run_generation_tab3, args=(ctx,))
        thread.daemon = True
        thread.start()

    def run_generation_tab3(self, ctx):
        """Run PDF generation in a background thread.

        Uses only the snapshot *ctx* dict — never reads tkinter
        StringVars or mutable instance state directly.
        """
        try:
            # Determine output folder: use custom dir if set, else default
            if ctx['output_dir'] and os.path.isdir(ctx['output_dir']):
                output_folder = ctx['output_dir']
            else:
                excel_dir = os.path.dirname(ctx['excel_path'])
                output_folder = os.path.join(excel_dir, "Completed Applications")
            os.makedirs(output_folder, exist_ok=True)

            total = len(ctx['selected_indices'])
            success_count = 0
            error_count = 0

            for i, idx in enumerate(ctx['selected_indices']):
                row = ctx['df'].iloc[idx]
                row_dict = {str(col).lower(): val for col, val in row.items()}

                # Get name for filename
                first_name = str(row_dict.get('first name', 'Unknown')).strip()
                surname = str(row_dict.get('surname', 'Unknown')).strip()

                if first_name.lower() == 'nan':
                    first_name = 'Unknown'
                if surname.lower() == 'nan':
                    surname = 'Unknown'

                # Clean names for filename (remove invalid characters)
                safe_first = "".join(c for c in first_name if c.isalnum() or c in ' -_').strip()
                safe_surname = "".join(c for c in surname if c.isalnum() or c in ' -_').strip()
                safe_school = "".join(c for c in ctx['school_name'] if c.isalnum() or c in ' -_').strip()
                safe_year = "".join(c for c in ctx['school_year'] if c.isalnum() or c in ' -_').strip()

                # Create filename
                filename = f"{safe_first}_{safe_surname}_Evidence Application {safe_school} {safe_year}.pdf"
                output_path = os.path.join(output_folder, filename)

                # Avoid overwriting existing files (e.g. duplicate names, re-runs)
                if os.path.exists(output_path):
                    base, ext = os.path.splitext(output_path)
                    counter = 1
                    while os.path.exists(f"{base} ({counter}){ext}"):
                        counter += 1
                    output_path = f"{base} ({counter}){ext}"

                try:
                    self._generate_single_pdf(ctx, row, output_path)
                    success_count += 1
                    status_text = f"Created: {filename}"
                except Exception as e:
                    error_count += 1
                    status_text = f"Error: {surname} - {str(e)}"

                # Update progress (UI calls are safe via root.after)
                progress = ((i + 1) / total) * 100
                self.root.after(0, self.update_progress_tab3, progress, status_text, i+1, total)

            # Final message
            final_message = f"Complete! {success_count} PDFs created"
            if error_count > 0:
                final_message += f", {error_count} errors"
            final_message += f"\n\nOutput folder: {output_folder}"

            self.root.after(0, self.generation_complete_tab3, final_message, output_folder)

        except Exception as e:
            err_msg = str(e)  # Capture before 'e' goes out of scope (PEP 3110)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Error", f"Generation failed:\n{msg}"))
            self.root.after(0, lambda: self.generate_btn_tab3.config(state=tk.NORMAL))

    def _generate_single_pdf(self, ctx, row_data, output_path):
        """Generate a single PDF with combed field support.

        Uses only the snapshot *ctx* dict for configuration and field
        metadata — safe to call from any thread.
        """
        reader = PdfReader(ctx['pdf_path'])
        writer = PdfWriter()

        # Clone the PDF
        writer.append(reader)

        # Create a dictionary of field values to fill
        field_values = {}

        # Check if we have analyzed fields with combed field metadata
        if ctx['analyzed_fields']:
            # Use combed field filler for smart filling
            combed_filler = CombedFieldFiller(settings={
                'padding': ctx['combed_padding'],
                'align': ctx['combed_align']
            })

            # Map Excel columns to values (case-insensitive)
            row_dict_lower = {str(col).lower(): self.format_value_tab3(val)
                             for col, val in row_data.items()}

            # Process each analyzed field
            for field in ctx['analyzed_fields']:
                # Find matching Excel column
                field_name_lower = field.field_name.lower()
                value = None

                for col_name, col_value in row_dict_lower.items():
                    if col_name == field_name_lower:
                        value = col_value
                        break

                if value is None:
                    continue

                # Fill field (handles both combed and regular)
                filled_values = combed_filler.fill_field(field, value)
                field_values.update(filled_values)

        else:
            # Fallback to original auto-matching (no combed field support)
            row_dict_lower = {str(col).lower(): val for col, val in row_data.items()}

            for pdf_field in ctx['pdf_fields']:
                pdf_field_lower = pdf_field.lower()

                # Try to find matching Excel column
                if pdf_field_lower in row_dict_lower:
                    val = self.format_value_tab3(row_dict_lower[pdf_field_lower])
                    field_values[pdf_field] = val

        # Fill all pages
        for page in writer.pages:
            writer.update_page_form_field_values(page, field_values)

        # Save the filled PDF
        with open(output_path, 'wb') as f:
            writer.write(f)

    @staticmethod
    def format_value_tab3(val):
        """Format a value for PDF insertion."""
        if pd.isna(val):
            return ""

        # Handle datetime objects (Australian format)
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.strftime('%d/%m/%Y')

        # Convert to string and clean
        str_val = str(val).strip()
        if str_val.lower() == 'nan':
            return ""

        return str_val

    def update_progress_tab3(self, progress, status, current, total):
        """Update progress for Tab 3."""
        self.progress_var_tab3.set(progress)
        self.progress_label_tab3.config(text=f"[{current}/{total}] {status}")

    def generation_complete_tab3(self, message, output_folder):
        """Handle generation completion for Tab 3."""
        self.progress_label_tab3.config(text="Generation complete!")
        self.update_status("Generation complete!", 'success')
        self.update_selection_count_tab3()  # Re-enable button with correct state

        # Ask to open folder
        result = messagebox.askyesno(
            "Generation Complete",
            f"{message}\n\nWould you like to open the output folder?"
        )

        if result:
            # Open folder in system file manager
            try:
                if sys.platform == 'darwin':
                    import subprocess
                    subprocess.run(['open', output_folder], check=False)
                elif sys.platform == 'win32':
                    os.startfile(output_folder)
                else:
                    import subprocess
                    subprocess.run(['xdg-open', output_folder], check=False)
            except Exception:
                pass  # Failing to open folder is non-critical


def main():
    root = tk.Tk()

    # Force window to front on launch
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))

    # Apply dark theme
    apply_dark_theme(root)

    app = VCAAPDFGeneratorV2(root)
    root.mainloop()


if __name__ == "__main__":
    main()
