#!/usr/bin/env python3
"""
Bulk PDF Generator v2.5
A GUI application to analyze PDF templates, map fields, and batch-fill forms.

Features:
- Multi-stage tabbed interface with modern dark theme
- PDF template analysis with visual field preview
- Combed field detection and smart filling
- Field mapping configuration
- Template library management
- Batch PDF generation from Excel data
"""

import os
import sys
import webbrowser
import tkinter as tk
import ttkbootstrap as ttk
from tkinter import filedialog, messagebox, simpledialog
from datetime import datetime
import pandas as pd
from pypdf import PdfReader, PdfWriter
import copy
import threading
import json
from pathlib import Path
from typing import List, Optional, Dict
from io import BytesIO
from PIL import Image, ImageTk, ImageDraw, ImageFont

def get_resource_path(filename: str) -> str:
    """Return the absolute path to a bundled resource file.

    Works correctly in both normal Python execution and when frozen as a
    PyInstaller single-file executable (where files are extracted to a
    temporary directory referenced by sys._MEIPASS at runtime).
    """
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


def _get_build_info() -> tuple:
    """Return (commit_hash, build_date, version_tag) for display.

    Priority:
    1. _version.py — baked in at build time by _generate_version.py (works in frozen exe)
    2. git subprocess — live read when running from source
    3. Fallback strings if neither is available
    """
    try:
        import _version
        commit = _version.BUILD_COMMIT
        date = _version.BUILD_DATE
        version = getattr(_version, 'BUILD_VERSION', 'dev')
        return (commit, date, version)
    except ImportError:
        pass

    try:
        import subprocess
        commit = subprocess.check_output(
            ['git', 'log', '-1', '--format=%h'],
            stderr=subprocess.DEVNULL, text=True
        ).strip()
        date = subprocess.check_output(
            ['git', 'log', '-1', '--format=%cd', '--date=format:%d %b %Y'],
            stderr=subprocess.DEVNULL, text=True
        ).strip()
        try:
            version = subprocess.check_output(
                ['git', 'describe', '--tags', '--abbrev=0'],
                stderr=subprocess.DEVNULL, text=True
            ).strip()
        except Exception:
            version = 'dev'
        return (commit, date, version)
    except Exception:
        pass

    return ('dev', 'local build', 'dev')


def check_for_update(current_version: str) -> dict:
    """Query GitHub Releases API and compare to current_version.

    Args:
        current_version: Version string like 'v2.6'. Pass 'dev' to skip check.

    Returns:
        dict with keys:
          'status'   — 'update_available' | 'up_to_date' | 'error'
          'latest'   — tag string from GitHub (present on non-error)
          'html_url' — release page URL (present on non-error)
          'message'  — human-readable error text (present on 'error' only)
    """
    import ssl
    import urllib.request
    import certifi

    RELEASES_API = (
        'https://api.github.com/repos/'
        'mrdavearms/bulk-pdf-extractor-and-generator/releases/latest'
    )

    def _parse_version(tag: str):
        """Convert 'v2.6' -> (2, 6) for numeric comparison."""
        try:
            return tuple(int(x) for x in tag.lstrip('v').split('.'))
        except ValueError:
            return (0,)

    # Don't prompt dev/source-run users — they have no installed version to update
    if not current_version.startswith('v'):
        return {'status': 'up_to_date', 'latest': current_version, 'html_url': ''}

    try:
        ctx = ssl.create_default_context(cafile=certifi.where())
        req = urllib.request.Request(
            RELEASES_API,
            headers={'User-Agent': 'BulkPDFGenerator-UpdateCheck/1.0'}
        )
        with urllib.request.urlopen(req, timeout=10, context=ctx) as resp:
            data = json.loads(resp.read())

        latest_tag = data.get('tag_name', '')
        html_url = data.get('html_url', '')
        is_newer = _parse_version(latest_tag) > _parse_version(current_version)

        return {
            'status': 'update_available' if is_newer else 'up_to_date',
            'latest': latest_tag,
            'html_url': html_url,
        }

    except Exception as exc:
        return {
            'status': 'error',
            'message': str(exc),
        }


def _resolve_data_dir() -> str:
    """Return the app data directory, creating it if needed.

    Tries ~/Documents/BulkPDFGenerator first. If that fails (common on school
    networks where Documents is a network-redirected folder that isn't mounted
    at login time), falls back to %LOCALAPPDATA%/BulkPDFGenerator on Windows,
    or ~/BulkPDFGenerator on other platforms. LOCALAPPDATA is always a local
    path and cannot be redirected by Group Policy.
    """
    primary = os.path.expanduser("~/Documents/BulkPDFGenerator")
    try:
        os.makedirs(primary, exist_ok=True)
        return primary
    except OSError:
        fallback = os.path.join(
            os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
            'BulkPDFGenerator'
        )
        os.makedirs(fallback, exist_ok=True)
        return fallback


# Import our new modules
from models import PDFField, TemplateConfig, AppSettings
from pdf_analyzer import PDFAnalyzer, auto_name_template
from visual_preview import VisualPreviewGenerator
from combed_filler import CombedFieldFiller
from theme import (
    COLORS, SPACING, SYSTEM_FONTS, font,
    apply_dark_theme, setup_treeview_tags, bind_treeview_hover,
)
from markdown_renderer import load_and_render


class ScrollableFrame(ttk.Frame):
    """A helper class to create a scrollable frame with dark-themed canvas.

    Mousewheel scrolling is scoped: only the frame under the cursor
    scrolls, and the delta calculation is platform-aware (Windows,
    macOS, Linux all behave differently).
    """

    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(
            self,
            borderwidth=0,
            highlightthickness=0,
            bg=COLORS['bg_base'],
            autostyle=False,
        )
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Scoped mousewheel: only bind when cursor is over this frame
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)
        self.scrollable_frame.bind("<Enter>", self._bind_mousewheel)
        self.scrollable_frame.bind("<Leave>", self._unbind_mousewheel)

    def _bind_mousewheel(self, event=None):
        """Bind mousewheel events to this specific canvas (not bind_all)."""
        if sys.platform == 'linux':
            self.canvas.bind("<Button-4>", self._on_mousewheel)
            self.canvas.bind("<Button-5>", self._on_mousewheel)
        else:
            self.canvas.bind("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, event=None):
        """Unbind mousewheel events when cursor leaves."""
        if sys.platform == 'linux':
            self.canvas.unbind("<Button-4>")
            self.canvas.unbind("<Button-5>")
        else:
            self.canvas.unbind("<MouseWheel>")

    def _on_mousewheel(self, event):
        """Scroll the canvas. Delta handling is platform-aware."""
        if not self.canvas.winfo_exists():
            return
        if sys.platform == 'darwin':
            # macOS: delta is already ±1 (or small integers)
            self.canvas.yview_scroll(int(-1 * event.delta), "units")
        elif sys.platform == 'linux':
            # Linux: Button-4 = scroll up, Button-5 = scroll down
            direction = -1 if event.num == 4 else 1
            self.canvas.yview_scroll(direction * 3, "units")
        else:
            # Windows: delta is ±120 per notch
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")



class SchoolSetupDialog(tk.Toplevel):
    """One-time dialog to capture school name and academic year."""

    def __init__(self, parent, current_name: str = "", current_year: str = ""):
        super().__init__(parent)
        self.title("School Setup")
        self.configure(bg=COLORS['bg_elevated'])
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.result_name = None
        self.result_year = None
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="School Details",
            font=(ff, 18, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            autostyle=False,
        ).pack(pady=(30, 4))

        tk.Label(
            self,
            text="These details are used in generated PDF filenames\n"
                 "and are saved so you only need to enter them once.",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
            justify=tk.CENTER,
            autostyle=False,
        ).pack(pady=(0, 20))

        # Form area
        form = tk.Frame(self, bg=C['bg_elevated'], autostyle=False)
        form.pack(padx=40, fill=tk.X)

        # School name
        tk.Label(
            form,
            text="School Name",
            font=(ff, 10, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
            autostyle=False,
        ).pack(fill=tk.X, pady=(0, 4))

        self.name_var = tk.StringVar(value=current_name)
        name_entry = ttk.Entry(form, textvariable=self.name_var, width=45)
        name_entry.pack(fill=tk.X, pady=(0, 4))
        name_entry.focus_set()

        tk.Label(
            form,
            text='e.g. "Wangaratta High School"',
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
            autostyle=False,
        ).pack(fill=tk.X, pady=(0, 16))

        # Academic year
        tk.Label(
            form,
            text="Academic Year",
            font=(ff, 10, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            anchor=tk.W,
            autostyle=False,
        ).pack(fill=tk.X, pady=(0, 4))

        # Default to current year if not provided
        if not current_year:
            current_year = str(datetime.now().year)
        self.year_var = tk.StringVar(value=current_year)
        year_entry = ttk.Entry(form, textvariable=self.year_var, width=10)
        year_entry.pack(anchor=tk.W, pady=(0, 16))

        # Save button
        ttk.Button(
            self,
            text="Save",
            command=self.on_save,
            bootstyle='primary',
        ).pack(pady=(8, 25))

        # Bind Enter key
        self.bind('<Return>', lambda e: self.on_save())

        # Auto-size and centre on parent
        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def on_save(self):
        name = self.name_var.get().strip()
        year = self.year_var.get().strip()
        if not name:
            messagebox.showwarning("Required", "Please enter your school name.",
                                   parent=self)
            return
        if not year:
            messagebox.showwarning("Required", "Please enter the academic year.",
                                   parent=self)
            return
        self.result_name = name
        self.result_year = year
        self.destroy()


class TemplateNameDialog(tk.Toplevel):
    """Dialog for naming a template before analysis."""

    def __init__(self, parent, suggested_name: str):
        super().__init__(parent)
        self.title("Template Name")
        self.geometry("520x280")
        self.configure(bg=COLORS['bg_elevated'])
        self.transient(parent)
        self.grab_set()

        self.result = None
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="Template Name",
            font=(ff, 16, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            autostyle=False,
        ).pack(pady=(20, 5))

        tk.Label(
            self,
            text="This configuration will be saved for reuse.",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
            autostyle=False,
        ).pack(pady=(0, 15))

        # Name entry
        name_frame = tk.Frame(self, bg=C['bg_elevated'], autostyle=False)
        name_frame.pack(pady=10, padx=40, fill=tk.X)

        tk.Label(name_frame, text="Template Name:",
                 font=(ff, 11), fg=C['text_primary'],
                 bg=C['bg_elevated'], autostyle=False).pack(anchor=tk.W, pady=(0, 5))

        self.name_var = tk.StringVar(value=suggested_name)
        ttk.Entry(name_frame, textvariable=self.name_var, width=50).pack(fill=tk.X)

        # Auto-generated info
        tk.Label(
            name_frame,
            text="Auto-generated from the PDF filename. Edit above to customise.",
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_elevated'],
            autostyle=False,
        ).pack(anchor=tk.W, pady=(5, 10))

        # Naming option
        self.naming_var = tk.StringVar(value="custom")

        ttk.Radiobutton(
            name_frame,
            text="Use PDF filename (auto-clean)",
            variable=self.naming_var,
            value="auto",
            style='Elevated.TRadiobutton',
        ).pack(anchor=tk.W, pady=2)

        ttk.Radiobutton(
            name_frame,
            text="Custom name (editable above)",
            variable=self.naming_var,
            value="custom",
            style='Elevated.TRadiobutton',
        ).pack(anchor=tk.W, pady=2)

        # Buttons
        button_frame = tk.Frame(self, bg=C['bg_elevated'], autostyle=False)
        button_frame.pack(pady=20, fill=tk.X, padx=40)

        ttk.Button(
            button_frame,
            text="Cancel",
            command=self.on_cancel,
        ).pack(side=tk.LEFT)

        ttk.Button(
            button_frame,
            text="Analyze & Save",
            command=self.on_save,
            bootstyle='primary',
        ).pack(side=tk.RIGHT)

        # Auto-size to content and center on parent
        self.update_idletasks()
        width = self.winfo_reqwidth()
        height = self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        # Force to front
        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def on_save(self):
        self.result = self.name_var.get().strip()
        if not self.result:
            messagebox.showwarning("Invalid Name", "Please enter a template name.")
            return
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()


# Date-hint keywords used to auto-detect date fields
_DATE_KEYWORDS = {'date', 'dob', 'birth', 'born', 'expiry', 'issued', 'due'}


def _guess_data_type(field_name: str) -> str:
    """Guess a field's data type from its name. Returns 'text', 'number', or 'date'."""
    lower = field_name.lower().replace('_', ' ')
    for kw in _DATE_KEYWORDS:
        if kw in lower:
            return 'date'
    return 'text'


class FieldTypeAuditDialog(tk.Toplevel):
    """Dialog for reviewing and setting field types and data types."""

    DATA_TYPE_OPTIONS = ['Text', 'Number', 'Date (DD/MM/YYYY)']
    _LABEL_TO_VALUE = {'Text': 'text', 'Number': 'number', 'Date (DD/MM/YYYY)': 'date'}
    _VALUE_TO_LABEL = {v: k for k, v in _LABEL_TO_VALUE.items()}

    FIELD_TYPE_OPTIONS = ['Text', 'Text-Combed']
    _EDITABLE_FIELD_TYPES = {'Text', 'Text-Combed'}  # Types the user can toggle between

    def __init__(self, parent, fields: list, preconfigured: set = None):
        super().__init__(parent)
        self.title("Review Field Types")
        self.configure(bg=COLORS['bg_elevated'])
        self.transient(parent)
        self.grab_set()

        self.fields = fields
        self.preconfigured = preconfigured or set()
        self.result = None  # Will be list of dicts on OK
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Title
        tk.Label(
            self,
            text="Review Field Types",
            font=(ff, 16, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_elevated'],
            autostyle=False,
        ).pack(pady=(20, 5))

        tk.Label(
            self,
            text="Set each field's type and data format.\n"
                 "Text-Combed fields require a character length.",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_elevated'],
            autostyle=False,
            justify=tk.CENTER,
        ).pack(pady=(0, 15))

        # Scrollable frame for field list
        list_outer = tk.Frame(self, bg=C['bg_elevated'], autostyle=False)
        list_outer.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 10))

        canvas = tk.Canvas(list_outer, bg=C['bg_surface'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_outer, orient=tk.VERTICAL, command=canvas.yview)
        self.inner_frame = tk.Frame(canvas, bg=C['bg_surface'], autostyle=False)

        self.inner_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=self.inner_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Mousewheel scrolling (platform-aware)
        def _on_mousewheel(event):
            if sys.platform == 'darwin':
                canvas.yview_scroll(int(-1 * event.delta), 'units')
            elif sys.platform == 'linux':
                direction = -1 if event.num == 4 else 1
                canvas.yview_scroll(direction * 3, 'units')
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')

        def _bind_scroll(widget):
            """Bind mousewheel events to a widget (cross-platform)."""
            widget.bind('<MouseWheel>', _on_mousewheel)
            if sys.platform == 'linux':
                widget.bind('<Button-4>', _on_mousewheel)
                widget.bind('<Button-5>', _on_mousewheel)

        _bind_scroll(canvas)
        _bind_scroll(self.inner_frame)

        # Header row
        hdr = tk.Frame(self.inner_frame, bg=C['bg_surface'], autostyle=False)
        hdr.pack(fill=tk.X, padx=5, pady=(5, 2))
        _bind_scroll(hdr)
        for hdr_text, hdr_w in [("Field Name", 26), ("Field Type", 14), ("Data Type", 16), ("Length", 8)]:
            lbl = tk.Label(hdr, text=hdr_text, font=(ff, 10, 'bold'),
                           fg=C['text_primary'], bg=C['bg_surface'], width=hdr_w,
                           anchor=tk.W, autostyle=False)
            lbl.pack(side=tk.LEFT)
            _bind_scroll(lbl)

        # Field rows
        self.combo_vars = []       # Data type combobox StringVars
        self.ftype_vars = []       # Field type combobox StringVars
        self.length_entries = []   # Length (Entry widget, StringVar) tuples

        for i, field in enumerate(fields):
            bg = C['bg_surface'] if i % 2 == 0 else C['bg_elevated']
            row = tk.Frame(self.inner_frame, bg=bg, autostyle=False)
            row.pack(fill=tk.X, padx=5, pady=1)
            _bind_scroll(row)

            # Column 1: Field Name
            lbl_name = tk.Label(row, text=field.field_name, font=(ff, 10),
                                fg=C['text_primary'], bg=bg, width=26,
                                anchor=tk.W, autostyle=False)
            lbl_name.pack(side=tk.LEFT)
            _bind_scroll(lbl_name)

            # Column 2: Field Type (editable for Text types, read-only for others)
            if field.field_type in self._EDITABLE_FIELD_TYPES:
                ftype_var = tk.StringVar(value=field.field_type)
                ftype_combo = ttk.Combobox(row, textvariable=ftype_var,
                                            values=self.FIELD_TYPE_OPTIONS,
                                            state='readonly', width=14)
                ftype_combo.pack(side=tk.LEFT, padx=(2, 0))
                _bind_scroll(ftype_combo)
                # Bind change handler (closure over index)
                ftype_combo.bind('<<ComboboxSelected>>',
                                 lambda e, idx=i: self._on_field_type_changed(idx))
            else:
                ftype_var = tk.StringVar(value=field.field_type)
                lbl_ftype = tk.Label(row, text=field.field_type, font=(ff, 10),
                                     fg=C['text_secondary'], bg=bg, width=14,
                                     anchor=tk.W, autostyle=False)
                lbl_ftype.pack(side=tk.LEFT, padx=(2, 0))
                _bind_scroll(lbl_ftype)
            self.ftype_vars.append(ftype_var)

            # Column 3: Data Type
            if field.field_name in self.preconfigured:
                default_type = field.data_type
            elif field.data_type != 'text':
                default_type = field.data_type
            else:
                default_type = _guess_data_type(field.field_name)
            dtype_var = tk.StringVar(value=self._VALUE_TO_LABEL.get(default_type, 'Text'))
            combo = ttk.Combobox(row, textvariable=dtype_var,
                                 values=self.DATA_TYPE_OPTIONS,
                                 state='readonly', width=16)
            combo.pack(side=tk.LEFT, padx=(5, 0))
            _bind_scroll(combo)
            self.combo_vars.append(dtype_var)

            # Column 4: Length (enabled only for Text-Combed)
            length_var = tk.StringVar(value=str(field.length) if field.length else "")
            length_entry = ttk.Entry(row, textvariable=length_var, width=8)
            length_entry.pack(side=tk.LEFT, padx=(5, 0))
            if field.field_type != 'Text-Combed':
                length_entry.config(state='disabled')
            self.length_entries.append((length_entry, length_var))

        # Buttons
        btn_frame = tk.Frame(self, bg=C['bg_elevated'], autostyle=False)
        btn_frame.pack(pady=15, fill=tk.X, padx=30)

        ttk.Button(btn_frame, text="Skip (all Text)",
                   command=self.on_skip).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Apply",
                   command=self.on_apply,
                   bootstyle='primary').pack(side=tk.RIGHT)

        # Size and center — use most of the parent height so buttons
        # remain visible even with many fields (the list scrolls).
        self.update_idletasks()
        dialog_w = max(740, self.winfo_reqwidth())
        parent_h = parent.winfo_height()
        dialog_h = min(int(parent_h * 0.85), max(500, 200 + len(fields) * 28))
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (dialog_w // 2)
        y = parent.winfo_y() + (parent_h // 2) - (dialog_h // 2)
        self.geometry(f"{dialog_w}x{dialog_h}+{x}+{y}")

        self.lift()
        self.attributes('-topmost', True)
        self.after(100, lambda: self.attributes('-topmost', False))

    def _on_field_type_changed(self, idx):
        """Enable/disable Length entry when field type changes."""
        new_type = self.ftype_vars[idx].get()
        entry, var = self.length_entries[idx]
        if new_type == 'Text-Combed':
            entry.config(state='normal')
            entry.focus_set()
        else:
            var.set("")
            entry.config(state='disabled')

    def on_apply(self):
        # Validate: any Text-Combed field must have a valid length
        for i, field in enumerate(self.fields):
            ftype = self.ftype_vars[i].get()
            if ftype == 'Text-Combed':
                entry, var = self.length_entries[i]
                length_str = var.get().strip()
                if not length_str or not length_str.isdigit() or int(length_str) < 1:
                    messagebox.showwarning(
                        "Invalid Length",
                        f"Field '{field.field_name}' is set to Text-Combed "
                        f"but has no valid character length.\n\n"
                        f"Please enter a positive number.",
                        parent=self,
                    )
                    entry.focus_set()
                    return

        # Collect results
        self.result = []
        for i in range(len(self.fields)):
            dtype = self._LABEL_TO_VALUE[self.combo_vars[i].get()]
            ftype = self.ftype_vars[i].get()
            _, length_var = self.length_entries[i]
            length_str = length_var.get().strip()
            length = int(length_str) if length_str and length_str.isdigit() else None

            self.result.append({
                'data_type': dtype,
                'field_type': ftype,
                'length': length,
            })
        self.destroy()

    def on_skip(self):
        self.result = None
        self.destroy()


class BulkPDFGenerator:
    """Main application class with tabbed interface."""

    def __init__(self, root):
        self.root = root
        self._build_info = _get_build_info()   # (commit, date, version) — cached once
        _commit, _date, _version_tag = self._build_info
        self.root.title(f"Bulk PDF Generator {_version_tag}")
        self.root.geometry("1000x800")
        self.root.minsize(900, 700)

        # Icon references — must live on self to prevent garbage collection
        self._icon_refs = {}
        self._load_app_icon()

        # Settings — resolve data dir with fallback for network-redirected Documents
        _data_dir = _resolve_data_dir()
        self.settings_file = os.path.join(_data_dir, 'settings.json')
        self.settings = self.load_settings()

        # If templates_directory still points to the Documents default, remap it to
        # the resolved data dir (covers the fallback case transparently).
        _docs_default = os.path.expanduser("~/Documents/BulkPDFGenerator/templates")
        if self.settings.templates_directory == _docs_default:
            self.settings.templates_directory = os.path.join(_data_dir, 'templates')

        # Ensure templates directory exists
        os.makedirs(self.settings.templates_directory, exist_ok=True)

        # Current state
        self.current_template: Optional[TemplateConfig] = None
        self.analyzed_fields: List[PDFField] = []
        self.pdf_template_path = tk.StringVar()
        self.excel_file_path = tk.StringVar()
        self.pdf_fields: List[str] = []

        # Tab 2 (Map Fields) state
        self._tab2_combos: Dict[str, tuple] = {}   # {field_name: (combobox, status_label)}
        self._tab2_mapping_frame = None             # rebuilt on each refresh
        self._tab2_status_label = None              # "X of Y fields mapped" label
        self._tab2_file_label = None                # shows current Excel file path
        self._tab2_auto_btn = None                  # "Auto-Map All" button
        self._tab2_clear_btn = None                 # "Clear All Mappings" button

        # Tab 3 (Generate) state - from original app
        self.df = None
        self.selected_rows = {}
        self.critical_fields = ['surname', 'first name', 'vcaa student number']
        self.output_dir_path = tk.StringVar()  # Optional custom output directory

        # Register cleanup on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.setup_ui()


    def _load_app_icon(self):
        """Load icon.png / icon.ico and apply to window and pre-scale for UI use.

        Pre-scaled ImageTk.PhotoImage objects are stored in self._icon_refs so
        they are never garbage-collected while the app is running:
          'header' → 32×32  used in the header bar
          'about'  → 72×72  used in the About tab card
        """
        ico_path = get_resource_path('icon.ico')
        png_path = get_resource_path('icon.png')

        # Set taskbar / title-bar icon (Windows .ico already set in main();
        # this handles macOS/Linux via .png)
        try:
            if sys.platform != 'win32' and os.path.exists(png_path):
                _img = Image.open(png_path).convert('RGBA')
                _photo = ImageTk.PhotoImage(_img.resize((64, 64), Image.LANCZOS))
                self.root.iconphoto(True, _photo)
                self._icon_refs['window'] = _photo
        except Exception:
            pass  # Fall back to default tkinter icon

        # Pre-scale for UI use
        if os.path.exists(png_path):
            try:
                _img = Image.open(png_path).convert('RGBA')
                for size, key in [(32, 'header'), (72, 'about')]:
                    _scaled = _img.resize((size, size), Image.LANCZOS)
                    self._icon_refs[key] = ImageTk.PhotoImage(_scaled)
            except Exception:
                pass  # UI falls back to diamond glyph

    def _close_preview_generator(self):
        """Safely close the preview generator if open."""
        if hasattr(self, 'preview_generator') and self.preview_generator:
            try:
                self.preview_generator.__exit__(None, None, None)
            except Exception:
                pass
            self.preview_generator = None

    def on_closing(self):
        """Clean up resources before closing."""
        self._close_preview_generator()
        self.root.destroy()

    def load_settings(self) -> AppSettings:
        """Load app settings or create defaults."""
        if os.path.exists(self.settings_file):
            return AppSettings.from_file(self.settings_file)
        else:
            # Create default settings
            settings = AppSettings.get_defaults()
            os.makedirs(os.path.dirname(self.settings_file), exist_ok=True)
            settings.save_to_file(self.settings_file)
            return settings

    def has_templates(self) -> bool:
        """Check if any templates exist."""
        if not os.path.exists(self.settings.templates_directory):
            return False
        templates = list(Path(self.settings.templates_directory).glob("*.json"))
        return len(templates) > 0

    def prompt_school_setup(self):
        """Show school setup dialog (first launch only)."""
        dialog = SchoolSetupDialog(
            self.root,
            current_name=self.settings.school_name,
            current_year=self.settings.school_year,
        )
        self.root.wait_window(dialog)

        if dialog.result_name:
            self.settings.school_name = dialog.result_name
            self.settings.school_year = dialog.result_year
            self.settings.save_to_file(self.settings_file)
            self.update_status(
                f"School set: {dialog.result_name} ({dialog.result_year})", 'success'
            )
            # Update header if it's showing
            if hasattr(self, 'header_school'):
                self.header_school.config(
                    text=f"{self.settings.school_name}  |  {self.settings.school_year}"
                )



    def setup_ui(self):
        """Create the main UI with tabbed interface."""
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Main container
        main_frame = ttk.Frame(self.root, padding="0")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ── Header Bar ──
        header = tk.Frame(main_frame, bg=C['bg_surface'], height=80, autostyle=False)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        title_area = tk.Frame(header, bg=C['bg_surface'], autostyle=False)
        title_area.pack(side=tk.LEFT, padx=24, pady=12)

        # Title row with app icon
        title_row = tk.Frame(title_area, bg=C['bg_surface'], autostyle=False)
        title_row.pack(anchor=tk.W)
        if 'header' in self._icon_refs:
            tk.Label(title_row,
                image=self._icon_refs['header'],
                bg=C['bg_surface'],
            ).pack(side=tk.LEFT, padx=(0, 8))
        else:
            tk.Label(title_row,
                text="\u25c6",
                font=(ff, 18),
                fg=C['accent'],
                bg=C['bg_surface'],
            ).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(title_row,
            text="Bulk PDF Generator",
            font=(ff, 22, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_surface'],
            autostyle=False,
        ).pack(side=tk.LEFT)

        tk.Label(title_area,
            text="Generate filled PDFs from spreadsheet data",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
            autostyle=False,
        ).pack(anchor=tk.W, padx=(30, 0))

        info_area = tk.Frame(header, bg=C['bg_surface'], autostyle=False)
        info_area.pack(side=tk.RIGHT, padx=24, pady=12)

        # School name (clickable to edit)
        if self.settings.school_configured:
            school_text = f"{self.settings.school_name}  |  {self.settings.school_year}"
        else:
            school_text = "Click to set school details"
        self.header_school = tk.Label(info_area,
            text=school_text,
            font=(ff, 10, 'bold'),
            fg=C['accent'],
            bg=C['bg_surface'],
            cursor="hand2",
            autostyle=False,
        )
        self.header_school.pack(anchor=tk.E)
        self.header_school.bind("<Button-1>", lambda e: self.prompt_school_setup())

        self.header_status = tk.Label(info_area,
            text="No template loaded",
            font=(ff, 10),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
            autostyle=False,
        )
        self.header_status.pack(anchor=tk.E)

        # Accent stripe divider
        tk.Frame(main_frame, bg=C['accent'], height=3, autostyle=False).pack(fill=tk.X)

        # ── Content area ──
        content_frame = ttk.Frame(main_frame, padding=str(SPACING['page_padding']))
        content_frame.pack(fill=tk.BOTH, expand=True)

        # Notebook (tabs)
        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Tab 0 uses a plain frame (Text widget has its own scrollbar)
        self.tab0_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab0_frame, text="  Getting Started  ")

        # Tabs 1-3 use ScrollableFrame
        self.tab1_container = ScrollableFrame(self.notebook)
        self.tab2_container = ScrollableFrame(self.notebook)
        self.tab3_container = ScrollableFrame(self.notebook)

        self.notebook.add(self.tab1_container, text="  1  Analyze Template  ")
        self.notebook.add(self.tab2_container, text="  2  Map Fields  ")
        self.notebook.add(self.tab3_container, text="  3  Generate PDFs  ")

        # About tab (plain frame, right-hand side)
        self.tab_about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_about_frame, text="  About  ")

        # Shorthand for actual UI frames
        self.tab1 = self.tab1_container.scrollable_frame
        self.tab2 = self.tab2_container.scrollable_frame
        self.tab3 = self.tab3_container.scrollable_frame

        # ── Status Bar ──
        tk.Frame(main_frame, bg=C['border_subtle'], height=1, autostyle=False).pack(fill=tk.X, side=tk.BOTTOM)
        status_frame = tk.Frame(main_frame, bg=C['bg_surface'], height=32, autostyle=False)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        status_frame.pack_propagate(False)

        self.status_label = tk.Label(status_frame,
            text="Ready",
            font=(ff, 9),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
            anchor=tk.W,
            autostyle=False,
        )
        self.status_label.pack(side=tk.LEFT, padx=12)

        self.status_template = tk.Label(status_frame,
            text="No template loaded",
            font=(ff, 9),
            fg=C['text_tertiary'],
            bg=C['bg_surface'],
            anchor=tk.E,
            autostyle=False,
        )
        self.status_template.pack(side=tk.RIGHT, padx=12)

        # Setup each tab
        self.setup_tab0_getting_started()
        self.setup_tab1_analyze()
        self.setup_tab2_mapping()
        self.setup_tab3_generate()
        self.setup_tab_about()

        # Tab 2 starts disabled; enabled once a PDF template is analyzed
        self.notebook.tab(2, state='disabled')

    # ========== UI HELPERS ==========

    def create_section(self, parent, title, subtitle=None, expand=False):
        """Create a card-style section with title label and bordered content area.

        Returns the inner content frame to pack widgets into.
        """
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Section wrapper
        wrapper = ttk.Frame(parent)
        fill_mode = tk.BOTH if expand else tk.X
        wrapper.pack(fill=fill_mode, expand=expand, pady=(0, SPACING['section_gap']))

        # Title label
        tk.Label(wrapper,
            text=title,
            font=(ff, 13, 'bold'),
            fg=C['text_primary'],
            bg=C['bg_base'],
            anchor=tk.W,
        ).pack(anchor=tk.W, pady=(0, 6))

        # Optional subtitle
        if subtitle:
            tk.Label(wrapper,
                text=subtitle,
                font=(ff, 10),
                fg=C['text_secondary'],
                bg=C['bg_base'],
                anchor=tk.W,
            ).pack(anchor=tk.W, pady=(0, 6))

        # Border frame (simulates 1px border via bg color + padding)
        border_frame = tk.Frame(wrapper, bg=C['border_subtle'], padx=1, pady=1,
                                autostyle=False)
        border_frame.pack(fill=fill_mode, expand=expand)

        # Inner content frame
        inner = tk.Frame(border_frame, bg=C['bg_surface'],
                         padx=SPACING['inner_padding'], pady=SPACING['inner_padding'],
                         autostyle=False)
        inner.pack(fill=fill_mode, expand=expand)

        return inner

    # ========== TAB 0: GETTING STARTED ==========

    def setup_tab0_getting_started(self):
        """Build the Getting Started tab with rendered markdown content."""
        C = COLORS

        # Text widget + scrollbar
        text_frame = ttk.Frame(self.tab0_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            bg=C['bg_base'],
            fg=C['text_primary'],
            insertbackground=C['text_primary'],
            selectbackground=C['accent_subtle'],
            selectforeground=C['text_primary'],
            borderwidth=0,
            highlightthickness=0,
            padx=20,
            pady=15,
            yscrollcommand=scrollbar.set,
            cursor='arrow',
            autostyle=False,
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)

        # Load and render the markdown file
        md_path = get_resource_path('getting_started.md')
        try:
            load_and_render(text_widget, md_path)
        except FileNotFoundError:
            text_widget.insert(tk.END, "Getting Started content not found.\n\n"
                               "Expected file: getting_started.md\n"
                               "Place it in the same folder as the application.")
            text_widget.config(state=tk.DISABLED)
        except (PermissionError, UnicodeDecodeError, OSError) as e:
            text_widget.insert(tk.END, f"Could not load Getting Started content:\n{e}")
            text_widget.config(state=tk.DISABLED)

    def setup_tab_about(self):
        """Build the About tab with developer info and project links."""
        C = COLORS
        ff = SYSTEM_FONTS['family']

        # Outer wrapper centres content vertically and horizontally
        outer = tk.Frame(self.tab_about_frame, bg=C['bg_base'])
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        # Centred card
        card_border = tk.Frame(outer, bg=C['border_subtle'], padx=1, pady=1,
                               autostyle=False)
        card_border.grid(row=0, column=0)  # centres in the expanded cell
        card = tk.Frame(card_border, bg=C['bg_surface'], padx=48, pady=40,
                        autostyle=False)
        card.pack()

        # App icon + title
        if 'about' in self._icon_refs:
            tk.Label(card, image=self._icon_refs['about'],
                     bg=C['bg_surface']).pack(pady=(0, 8))
        else:
            tk.Label(card, text="\u25c6", font=(ff, 36), fg=C['accent'],
                     bg=C['bg_surface']).pack(pady=(0, 4))
        tk.Label(card, text="Bulk PDF Generator", font=(ff, 22, 'bold'),
                 fg=C['text_primary'], bg=C['bg_surface']).pack()
        tk.Label(card, text="Generate filled PDFs from spreadsheet data",
                 font=(ff, 11), fg=C['text_secondary'],
                 bg=C['bg_surface']).pack(pady=(2, 20))

        # Divider
        tk.Frame(card, bg=C['border_subtle'], height=1).pack(fill=tk.X, pady=(0, 20))

        # Mission statement
        mission = (
            "Designed to take the pain out of complicated, repetitive\n"
            "PDF form-filling tasks in schools \u2014 turning hours of manual\n"
            "data entry into a single click, so staff can focus on the\n"
            "work that actually matters."
        )
        tk.Label(card, text=mission, font=(ff, 11), fg=C['text_primary'],
                 bg=C['bg_surface'], justify=tk.CENTER,
                 wraplength=420).pack(pady=(0, 24))

        # Developer info
        tk.Label(card, text="Developed by", font=(ff, 10),
                 fg=C['text_tertiary'], bg=C['bg_surface']).pack()
        tk.Label(card, text="Dave Armstrong", font=(ff, 14, 'bold'),
                 fg=C['text_primary'], bg=C['bg_surface']).pack(pady=(2, 2))
        tk.Label(card, text="A Principal-developed app for educators and school leaders",
                 font=(ff, 10), fg=C['text_secondary'],
                 bg=C['bg_surface']).pack(pady=(0, 16))

        # Email link
        email_label = tk.Label(card, text="Dave.Armstrong@education.vic.gov.au",
                               font=(ff, 11, 'underline'), fg=C['info'],
                               bg=C['bg_surface'], cursor='hand2')
        email_label.pack(pady=(0, 6))
        email_label.bind('<Button-1>',
                         lambda e: webbrowser.open('mailto:Dave.Armstrong@education.vic.gov.au'))

        # GitHub link
        github_label = tk.Label(card, text="github.com/mrdavearms/bulk-pdf-extractor-and-generator",
                                font=(ff, 11, 'underline'), fg=C['info'],
                                bg=C['bg_surface'], cursor='hand2')
        github_label.pack(pady=(0, 20))
        github_label.bind('<Button-1>',
                          lambda e: webbrowser.open('https://github.com/mrdavearms/bulk-pdf-extractor-and-generator'))

        # Divider
        tk.Frame(card, bg=C['border_subtle'], height=1).pack(fill=tk.X, pady=(0, 16))

        # Disclaimer
        disclaimer = (
            "This is a Principal-developed tool shared in good faith.\n"
            "Always review all generated outputs before use."
        )
        tk.Label(card, text=disclaimer, font=(ff, 9), fg=C['text_tertiary'],
                 bg=C['bg_surface'], justify=tk.CENTER).pack(pady=(0, 12))

        # Version — commit hash and build date baked in at build time
        _commit, _date, _version_tag = self._build_info
        version_str = f"{_version_tag}  ·  {_commit}  ·  {_date}"
        tk.Label(card, text=version_str, font=(ff, 10),
                 fg=C['text_tertiary'], bg=C['bg_surface']).pack(pady=(0, 8))

        update_btn = ttk.Button(card, text='Check for Updates', bootstyle='outline-secondary',
                                width=20)
        update_btn.config(command=lambda: self._run_update_check(update_btn))
        update_btn.pack()

    def _run_update_check(self, button):
        """Start a background update check. Disables the button while running."""
        button.config(state='disabled', text='Checking…')

        def _worker():
            _commit, _date, current_version = self._build_info
            result = check_for_update(current_version)
            self.root.after(0, lambda: self._show_update_result(result, button))

        threading.Thread(target=_worker, daemon=True).start()

    def _show_update_result(self, result, button):
        """Display update-check result. Re-enables the button when done."""
        button.config(state='normal', text='Check for Updates')

        status = result.get('status')

        if status == 'update_available':
            latest = result['latest']
            url = result['html_url']
            go = messagebox.askyesno(
                'Update Available',
                f'Version {latest} is available.\n\nOpen the Releases page to download?',
                icon='info'
            )
            if go:
                webbrowser.open(url)

        elif status == 'up_to_date':
            messagebox.showinfo('Up to Date', "You're running the latest version.")

        else:
            messagebox.showwarning(
                'Update Check Failed',
                f"Could not check for updates.\n\n{result.get('message', 'Unknown error')}"
            )

    def update_status(self, message, level='info'):
        """Update the status bar message."""
        color_map = {
            'info': COLORS['text_secondary'],
            'success': COLORS['success'],
            'warning': COLORS['warning'],
            'error': COLORS['error'],
        }
        self.status_label.config(
            text=message,
            fg=color_map.get(level, COLORS['text_secondary'])
        )

    # ========== TAB 1: ANALYZE TEMPLATE ==========

    def setup_tab1_analyze(self):
        """Setup Tab 1: PDF Template Analysis."""
        tab = self.tab1

        # Container with padding
        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        # Section: Load Template
        load_inner = self.create_section(container, "Load Template")

        # PDF selection row
        pdf_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        pdf_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(pdf_row, text="PDF Template:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(pdf_row, textvariable=self.pdf_template_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(pdf_row, text="Browse...", command=self.select_pdf_tab1, width=10).pack(side=tk.LEFT)

        # Template name row
        name_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        name_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(name_row, text="Template Name:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        self.template_name_var = tk.StringVar()
        ttk.Entry(name_row, textvariable=self.template_name_var, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)

        # Naming options
        naming_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        naming_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']), padx=(144, 0))
        self.naming_option_var = tk.StringVar(value="auto")
        ttk.Radiobutton(naming_row, text="Auto-name from PDF", variable=self.naming_option_var, value="auto", style='Surface.TRadiobutton').pack(side=tk.LEFT, padx=(0, 12))
        ttk.Radiobutton(naming_row, text="Custom name", variable=self.naming_option_var, value="custom", style='Surface.TRadiobutton').pack(side=tk.LEFT)

        # Recent templates dropdown
        recent_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        recent_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(recent_row, text="Recent Templates:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        self.recent_templates_var = tk.StringVar()
        self.recent_templates_combo = ttk.Combobox(recent_row, textvariable=self.recent_templates_var, state="readonly", width=47)
        self.recent_templates_combo.pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(recent_row, text="Load", command=self.load_recent_template, width=10).pack(side=tk.LEFT)

        self.populate_recent_templates()

        # Analyze button
        btn_row = tk.Frame(load_inner, bg=COLORS['bg_surface'])
        btn_row.pack(fill=tk.X, pady=(SPACING['element_gap'], 0))
        ttk.Button(btn_row, text="Analyze Fields", command=self.analyze_pdf_fields, bootstyle='primary').pack()

        # Analysis results section
        results_inner = self.create_section(container, "Analysis Results", expand=True)

        # Stats row
        self.stats_label = tk.Label(results_inner, text="No analysis performed yet",
                                    font=font(10), fg=COLORS['text_secondary'], bg=COLORS['bg_surface'])
        self.stats_label.pack(pady=(0, 8))

        # Fields table
        table_frame = tk.Frame(results_inner, bg=COLORS['bg_surface'])
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('field_name', 'type', 'page', 'length', 'data_type')
        self.fields_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

        self.fields_tree.heading('field_name', text='Field Name')
        self.fields_tree.heading('type', text='Type')
        self.fields_tree.heading('page', text='Page')
        self.fields_tree.heading('length', text='Length/Size')
        self.fields_tree.heading('data_type', text='Data Type')

        self.fields_tree.column('field_name', width=220)
        self.fields_tree.column('type', width=110)
        self.fields_tree.column('page', width=55)
        self.fields_tree.column('length', width=90)
        self.fields_tree.column('data_type', width=110)

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.fields_tree.yview)
        self.fields_tree.configure(yscrollcommand=scrollbar.set)

        self.fields_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Setup treeview tags and hover
        setup_treeview_tags(self.fields_tree)
        bind_treeview_hover(self.fields_tree)
        self.fields_tree.tag_configure('date_field', foreground='#d97706')  # Amber for date fields

        # Bind selection event for field preview
        self.fields_tree.bind('<<TreeviewSelect>>', self.on_field_selected)
        # Double-click to edit data type inline
        self.fields_tree.bind('<Double-1>', self._on_field_double_click)

        # Visual Preview section (nested inside results card)
        preview_header = tk.Frame(results_inner, bg=COLORS['bg_surface'])
        preview_header.pack(fill=tk.X, pady=(SPACING['element_gap'], 6))

        tk.Label(preview_header, text="Field Preview",
                 font=font(11, 'bold'), fg=COLORS['text_primary'],
                 bg=COLORS['bg_surface'], anchor=tk.W).pack(side=tk.LEFT)

        # Zoom controls
        zoom_frame = tk.Frame(preview_header, bg=COLORS['bg_surface'])
        zoom_frame.pack(side=tk.RIGHT)
        self.zoom_level = 1.0
        self._preview_raw_img = None  # Store unscaled PIL image for zoom

        ttk.Button(zoom_frame, text="\u2212", width=3,
                   command=lambda: self._zoom_preview(-0.25)).pack(side=tk.LEFT, padx=2)
        self.zoom_label = tk.Label(zoom_frame, text="100%", width=5,
                                   font=font(10), fg=COLORS['text_secondary'],
                                   bg=COLORS['bg_surface'], anchor=tk.CENTER)
        self.zoom_label.pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="+", width=3,
                   command=lambda: self._zoom_preview(0.25)).pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="Fit", width=4,
                   command=lambda: self._zoom_preview(0, fit=True)).pack(side=tk.LEFT, padx=(6, 0))

        tk.Label(results_inner, text="Click a field above to preview \u2022 Use +/\u2212 to zoom \u2022 Scroll to pan",
                 font=font(9), fg=COLORS['text_secondary'], bg=COLORS['bg_surface'], anchor=tk.W).pack(fill=tk.X, pady=(0, 6))

        # Preview canvas (scrollable for zoom)
        canvas_frame = tk.Frame(results_inner, bg=COLORS['canvas_border'])
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.preview_canvas = tk.Canvas(
            canvas_frame, width=600, height=300,
            bg=COLORS['canvas_bg'],
            highlightthickness=0,
            xscrollincrement=1, yscrollincrement=1,
            autostyle=False,
        )
        self.preview_canvas.pack(fill=tk.BOTH, expand=True)

        # Mouse wheel scrolling on canvas for panning when zoomed
        def _on_canvas_scroll(event):
            if self.zoom_level > 1.0:
                if sys.platform == 'darwin':
                    self.preview_canvas.yview_scroll(int(-1 * event.delta), "units")
                elif sys.platform == 'linux':
                    direction = -1 if event.num == 4 else 1
                    self.preview_canvas.yview_scroll(direction * 3, "units")
                else:
                    self.preview_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.preview_canvas.bind('<MouseWheel>', _on_canvas_scroll)
        if sys.platform == 'linux':
            self.preview_canvas.bind('<Button-4>', _on_canvas_scroll)
            self.preview_canvas.bind('<Button-5>', _on_canvas_scroll)

        self.preview_image = None
        self.preview_generator = None

        # Action buttons row
        action_frame = tk.Frame(container, bg=COLORS['bg_base'])
        action_frame.pack(fill=tk.X, pady=(0, SPACING['element_gap']))

        ttk.Button(action_frame, text="Export Mapping File (.xlsx)", command=self.export_mapping_file).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(action_frame, text="Save Template Config", command=self.save_template_config, bootstyle='primary').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(action_frame, text="Skip to Generate PDFs \u2192", command=lambda: self.notebook.select(3)).pack(side=tk.RIGHT)

    def select_pdf_tab1(self):
        """Select PDF file in Tab 1."""
        filepath = filedialog.askopenfilename(
            title="Select PDF Template",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filepath:
            self.pdf_template_path.set(filepath)
            # Auto-generate template name
            if self.naming_option_var.get() == "auto":
                suggested_name = auto_name_template(os.path.basename(filepath))
                self.template_name_var.set(suggested_name)

    def populate_recent_templates(self):
        """Populate recent templates dropdown."""
        templates_dir = Path(self.settings.templates_directory)
        if not templates_dir.exists():
            return

        template_files = list(templates_dir.glob("*.json"))
        template_names = [f.stem for f in template_files]

        self.recent_templates_combo['values'] = template_names

    def load_recent_template(self):
        """Load a template from the recent list."""
        template_name = self.recent_templates_var.get()
        if not template_name:
            return

        template_path = os.path.join(self.settings.templates_directory, f"{template_name}.json")
        if os.path.exists(template_path):
            self.load_template_config(template_path)

    def load_template_from_file(self):
        """Load template config from file dialog."""
        filepath = filedialog.askopenfilename(
            title="Select Template Configuration",
            initialdir=self.settings.templates_directory,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filepath:
            self.load_template_config(filepath)

    def load_template_config(self, filepath: str):
        """Load template configuration."""
        try:
            self.current_template = TemplateConfig.from_file(filepath)
            self.pdf_template_path.set(self.current_template.pdf_path)
            self.template_name_var.set(self.current_template.template_name)

            # Restore saved data types to analyzed fields if they exist
            saved_types = self.current_template.field_data_types or {}
            if saved_types and self.analyzed_fields:
                for field in self.analyzed_fields:
                    if field.field_name in saved_types:
                        field.data_type = saved_types[field.field_name]
                self.display_analyzed_fields()

            # Restore saved field→Excel column mappings
            saved_cols = self.current_template.field_excel_columns or {}
            if saved_cols and self.analyzed_fields:
                for field in self.analyzed_fields:
                    if field.field_name in saved_cols:
                        field.excel_column = saved_cols[field.field_name]
            self._refresh_tab2_mappings()

            self.update_status(f"Template loaded: {self.current_template.template_name}", 'success')
            self.header_status.config(text=f"Template: {self.current_template.template_name}")
            self.status_template.config(text=self.current_template.template_name)
            messagebox.showinfo("Template Loaded", f"Loaded template: {self.current_template.template_name}")

            # Switch to Tab 2 or 3
            self.notebook.select(3)  # Go to Generate tab

        except Exception as e:
            self.update_status(f"Failed to load template", 'error')
            messagebox.showerror("Error", f"Failed to load template:\n{str(e)}")

    def analyze_pdf_fields(self):
        """Analyze PDF fields and detect combed fields."""
        pdf_path = self.pdf_template_path.get()

        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "Please select a valid PDF template file.")
            return

        try:
            # Get template name (show dialog if custom)
            if self.naming_option_var.get() == "custom":
                suggested_name = self.template_name_var.get() or auto_name_template(os.path.basename(pdf_path))
                dialog = TemplateNameDialog(self.root, suggested_name)
                self.root.wait_window(dialog)

                if dialog.result is None:
                    return  # User cancelled

                template_name = dialog.result
            else:
                template_name = auto_name_template(os.path.basename(pdf_path))

            self.template_name_var.set(template_name)

            # Analyze PDF
            with PDFAnalyzer(pdf_path) as analyzer:
                self.analyzed_fields = analyzer.analyze_fields()
                field_stats = analyzer.get_field_statistics(self.analyzed_fields)

            # Close any existing preview generator before opening a new one
            self._close_preview_generator()

            # Initialize preview generator
            cache_dir = os.path.join(self.settings.templates_directory, '.preview_cache')
            self.preview_generator = VisualPreviewGenerator(pdf_path, cache_dir)
            self.preview_generator.__enter__()  # Open the PDF

            # Update UI
            self.display_analyzed_fields()

            # Update stats label
            total = len(self.analyzed_fields)
            combed = sum(1 for f in self.analyzed_fields if f.is_combed)
            pages = len(set(f.page for f in self.analyzed_fields))

            self.stats_label.config(
                text=f"Template: {template_name}  |  Total Fields: {total}  |  Pages: {pages}  |  Combed: {combed}"
            )

            # Update status bar and header
            self.update_status(f"Analysis complete: {total} fields found ({combed} combed)", 'success')
            self.header_status.config(text=f"Template: {template_name}")
            self.status_template.config(text=f"{template_name} ({total} fields)")

            messagebox.showinfo("Analysis Complete", f"Found {total} fields ({combed} combed fields)")

            # Restore saved overrides from current template if available
            preconfigured_fields = set()
            if self.current_template:
                # Restore data types
                if self.current_template.field_data_types:
                    for field in self.analyzed_fields:
                        saved = self.current_template.field_data_types.get(field.field_name)
                        if saved:
                            field.data_type = saved
                            preconfigured_fields.add(field.field_name)

                # Restore field type overrides
                if self.current_template.field_type_overrides:
                    for field in self.analyzed_fields:
                        override = self.current_template.field_type_overrides.get(field.field_name)
                        if override:
                            field.field_type = override.get('field_type', field.field_type)
                            if override.get('length') is not None:
                                field.length = override['length']
                                field.is_combed = (field.field_type == 'Text-Combed')
                            preconfigured_fields.add(field.field_name)

                # Restore field→Excel column mappings
                if self.current_template.field_excel_columns:
                    for field in self.analyzed_fields:
                        saved_col = self.current_template.field_excel_columns.get(field.field_name)
                        if saved_col:
                            field.excel_column = saved_col

            # Show field type audit dialog
            audit = FieldTypeAuditDialog(self.root, self.analyzed_fields, preconfigured_fields)
            self.root.wait_window(audit)
            if audit.result is not None:
                for field, res in zip(self.analyzed_fields, audit.result):
                    field.data_type = res['data_type']
                    field.field_type = res['field_type']
                    if res['field_type'] == 'Text-Combed':
                        field.is_combed = True
                        field.length = res['length']
                        # Keep combed_fields if already populated (sub-field combed);
                        # leave empty for single-field combed
                    else:
                        field.is_combed = False
                        field.length = None
                        field.combed_fields = []
                self.display_analyzed_fields()  # Refresh to show updated types

            # Enable Tab 2 and populate it (auto-map if Excel already loaded)
            self.notebook.tab(2, state='normal')
            self._auto_map_fields()
            self._refresh_tab2_mappings()

        except Exception as e:
            self._close_preview_generator()  # Guarantee cleanup on error
            self.update_status(f"Analysis failed: {str(e)}", 'error')
            messagebox.showerror("Error", f"Failed to analyze PDF:\n{str(e)}")

    def display_analyzed_fields(self):
        """Display analyzed fields in the table."""
        # Clear existing
        for item in self.fields_tree.get_children():
            self.fields_tree.delete(item)

        # Add fields with alternating row colors
        for i, field in enumerate(self.analyzed_fields):
            length_display = f"{field.length} chars" if field.is_combed else "Single"
            data_type_label = FieldTypeAuditDialog._VALUE_TO_LABEL.get(field.data_type, 'Text')
            row_tag = 'odd' if i % 2 else 'even'
            tags_list = [row_tag]
            if field.is_combed:
                tags_list.append('combed')
            if field.data_type == 'date':
                tags_list.append('date_field')

            self.fields_tree.insert('', tk.END, values=(
                field.field_name,
                field.field_type,
                field.page,
                length_display,
                data_type_label
            ), tags=tuple(tags_list))

    def _on_field_double_click(self, event):
        """Handle double-click on a field row to edit its data type inline."""
        region = self.fields_tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        col = self.fields_tree.identify_column(event.x)
        # col is '#5' for the 5th column (data_type)
        if col != '#5':
            return
        item = self.fields_tree.identify_row(event.y)
        if not item:
            return
        idx = self.fields_tree.index(item)
        if idx >= len(self.analyzed_fields):
            return

        field = self.analyzed_fields[idx]
        current_label = FieldTypeAuditDialog._VALUE_TO_LABEL.get(field.data_type, 'Text')

        # Get cell bounding box
        bbox = self.fields_tree.bbox(item, col)
        if not bbox:
            return

        # Create overlay combobox
        combo = ttk.Combobox(self.fields_tree,
                             values=FieldTypeAuditDialog.DATA_TYPE_OPTIONS,
                             state='readonly', width=16)
        combo.set(current_label)
        combo.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        combo.focus_set()
        combo.event_generate('<Button-1>')  # Open the dropdown

        _destroyed = [False]  # Mutable flag to guard against double-destroy

        def _safe_destroy():
            if not _destroyed[0]:
                _destroyed[0] = True
                try:
                    combo.destroy()
                except tk.TclError:
                    pass

        def _on_select(e=None):
            new_label = combo.get()
            new_value = FieldTypeAuditDialog._LABEL_TO_VALUE.get(new_label, 'text')
            field.data_type = new_value
            _safe_destroy()
            self.display_analyzed_fields()

        def _on_escape(e=None):
            _safe_destroy()

        combo.bind('<<ComboboxSelected>>', _on_select)
        combo.bind('<Escape>', _on_escape)
        combo.bind('<FocusOut>', lambda e: combo.after(100, _safe_destroy))

    def _zoom_preview(self, delta, fit=False):
        """Adjust preview zoom level and re-render."""
        if fit:
            self.zoom_level = 1.0
        else:
            self.zoom_level = max(0.5, min(4.0, self.zoom_level + delta))

        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")
        self._render_preview_at_zoom()

    def _render_preview_at_zoom(self):
        """Re-render the stored raw preview image at the current zoom level."""
        if self._preview_raw_img is None:
            return

        try:
            canvas_width = self.preview_canvas.winfo_width()
            canvas_height = self.preview_canvas.winfo_height()
            if canvas_width <= 1:
                canvas_width = 600
            if canvas_height <= 1:
                canvas_height = 300

            img_width, img_height = self._preview_raw_img.size

            # Base scale: fit to canvas
            base_scale = min(canvas_width / img_width, canvas_height / img_height) * 0.95
            scale = base_scale * self.zoom_level

            new_width = max(1, int(img_width * scale))
            new_height = max(1, int(img_height * scale))

            resized = self._preview_raw_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.preview_image = ImageTk.PhotoImage(resized)

            self.preview_canvas.delete("all")

            if self.zoom_level > 1.0:
                # Enable scrolling for zoomed images
                self.preview_canvas.config(scrollregion=(0, 0, new_width, new_height))
                self.preview_canvas.create_image(new_width // 2, new_height // 2,
                                                 image=self.preview_image, anchor=tk.CENTER)
            else:
                # Reset scroll region to canvas size when fitting
                self.preview_canvas.config(scrollregion=(0, 0, canvas_width, canvas_height))
                self.preview_canvas.create_image(canvas_width // 2, canvas_height // 2,
                                                 image=self.preview_image, anchor=tk.CENTER)
        except Exception as e:
            print(f"Zoom render error: {e}")

    def on_field_selected(self, event):
        """Handle field selection in Tab 1 - show visual preview."""
        if (not self.analyzed_fields
                or not self.preview_generator
                or not self.preview_generator.doc):
            return

        selection = self.fields_tree.selection()
        if not selection:
            return

        # Get selected item index
        item = selection[0]
        item_index = self.fields_tree.index(item)

        if item_index >= len(self.analyzed_fields):
            return

        # Get the selected field
        field = self.analyzed_fields[item_index]

        try:
            # Generate preview at higher DPI for zoom quality
            preview_img = self.preview_generator.generate_field_preview(field, dpi=200)

            # Store raw image for zoom
            self._preview_raw_img = preview_img
            self.zoom_level = 1.0
            self.zoom_label.config(text="100%")

            # Render at current zoom (handles canvas drawing)
            self._render_preview_at_zoom()

        except Exception as e:
            print(f"Preview error: {e}")
            self.preview_canvas.delete("all")
            self.preview_canvas.create_text(
                300, 150,
                text=f"Preview unavailable\n{str(e)}",
                fill=COLORS['error'],
                font=font(12),
            )

    def export_mapping_file(self):
        """Export field mapping to Excel file with formatted worksheets.

        Creates three sheets:
        - Field Mapping: detailed reference of all PDF fields
        - Data Entry: ready-to-use data entry sheet with field names as column headers
        - Instructions: how-to guide for using the file
        All cells have wrap-text formatting applied.
        """
        if not self.analyzed_fields:
            messagebox.showwarning("No Data", "Please analyze a PDF template first.")
            return

        template_name = self.template_name_var.get() or "template"
        suggested_filename = f"{template_name}_mapping.xlsx"

        filepath = filedialog.asksaveasfilename(
            title="Save Mapping File",
            initialdir=self.settings.templates_directory,
            initialfile=suggested_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not filepath:
            return

        try:
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            # Build field data
            data = []
            for field in self.analyzed_fields:
                excel_col = self.smart_guess_excel_column(field.field_name)
                data.append({
                    'PDF_Field_Name': field.field_name,
                    'Excel_Column_Name': excel_col,
                    'Field_Type': field.field_type,
                    'Page': field.page,
                    'Required': 'Yes' if field.field_name.lower() in [
                        'surname', 'first_name', 'first name',
                        'vcaa_number', 'vcaa student number'
                    ] else 'No',
                    'Length': f"{field.length} chars" if field.is_combed else '-',
                    'Notes': self.generate_field_notes(field),
                })

            df = pd.DataFrame(data)

            # Data Entry columns: use the Excel column names as headers
            data_entry_cols = [d['Excel_Column_Name'] for d in data if d['Excel_Column_Name']]
            df_data_entry = pd.DataFrame(columns=data_entry_cols)

            # Write sheets via pandas, then format with openpyxl
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Field Mapping', index=False)
                df_data_entry.to_excel(writer, sheet_name='Data Entry', index=False)

                # Instructions sheet
                instructions = pd.DataFrame({
                    'Bulk PDF Generator - Field Mapping Guide': [
                        '',
                        'HOW TO USE THIS FILE:',
                        '1. Review the "Field Mapping" sheet',
                        '2. Update any "Excel_Column_Name" values to match your preferred headers',
                        '3. The "Data Entry" sheet is where you can start typing student data',
                        '4. Save this file',
                        '5. Return to the app and load this file in Tab 3 (Generate PDFs)',
                        '',
                        'IMPORTANT NOTES:',
                        '- The "Data Entry" sheet headers are automatically generated',
                        '- If you change column names in "Field Mapping", you should also update them in "Data Entry"',
                        '- Combed fields will auto-split text (e.g., "John" → J-o-h-n)',
                        '- The app uses the "Data Entry" sheet for PDF generation',
                        '',
                        'For help: See the Getting Started tab in the app',
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False, header=False)

                wb = writer.book

                # -- Shared styles --
                wrap_align = Alignment(wrap_text=True, vertical='top')
                header_font = Font(bold=True, size=11)
                header_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                thin_border = Border(
                    bottom=Side(style='thin', color='E0E0E3'),
                )

                # ── Format: Field Mapping sheet ──
                ws_map = wb['Field Mapping']
                for row in ws_map.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_align
                # Bold + coloured header row
                for cell in ws_map[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                # Auto-width columns (min 14, max 40)
                for col in ws_map.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    adjusted = min(40, max(14, max_len + 4))
                    ws_map.column_dimensions[col[0].column_letter].width = adjusted

                # ── Format: Data Entry sheet ──
                ws_entry = wb['Data Entry']
                # Format all columns as Text so numbers stay as strings
                from openpyxl.utils import get_column_letter
                for c in range(1, len(data_entry_cols) + 1):
                    col_letter = get_column_letter(c)
                    for row_num in range(2, 502):  # rows 2-501 for data
                        ws_entry.cell(row=row_num, column=c).number_format = '@'
                # Apply wrap text to ALL cells (headers + 50 empty rows for data entry)
                for r in range(1, 52):
                    for c in range(1, len(data_entry_cols) + 1):
                        cell = ws_entry.cell(row=r, column=c)
                        cell.alignment = wrap_align
                # Bold + coloured header row with vertical centering
                header_align = Alignment(wrap_text=True, vertical='center')
                for cell in ws_entry[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_align
                # Expand header row height so wrapped text shows on ~2 lines
                ws_entry.row_dimensions[1].height = 36
                # Set column widths (min 16 based on header text)
                for idx, col_name in enumerate(data_entry_cols, 1):
                    width = min(30, max(16, len(col_name) + 4))
                    ws_entry.column_dimensions[ws_entry.cell(row=1, column=idx).column_letter].width = width
                # Freeze header row so it stays visible during scrolling
                ws_entry.freeze_panes = 'A2'

                # ── Format: Instructions sheet ──
                ws_instr = wb['Instructions']
                for row in ws_instr.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_align
                ws_instr.column_dimensions['A'].width = 80

                # ── About sheet ──
                _commit, _date, _version_tag = self._build_info
                ws_about = wb.create_sheet(title='About')

                about_rows = [
                    ('Bulk PDF Generator', 'title'),
                    ('Generate filled PDFs from spreadsheet data', 'subtitle'),
                    ('', None),
                    ('PURPOSE', 'heading'),
                    (
                        'Designed to take the pain out of complicated, repetitive PDF form-filling '
                        'tasks in schools — turning hours of manual data entry into a single click, '
                        'so staff can focus on the work that actually matters.',
                        'body'
                    ),
                    ('', None),
                    ('DEVELOPER', 'heading'),
                    ('Dave Armstrong', 'name'),
                    ('A Principal-developed app for educators and school leaders', 'body'),
                    ('', None),
                    ('CONTACT', 'heading'),
                    ('Email:   Dave.Armstrong@education.vic.gov.au', 'body'),
                    ('GitHub:  github.com/mrdavearms/bulk-pdf-extractor-and-generator', 'body'),
                    ('', None),
                    ('DISCLAIMER', 'heading'),
                    (
                        'This is a Principal-developed tool shared in good faith. '
                        'Always review all generated outputs before use.',
                        'body'
                    ),
                    ('', None),
                    (f'Version: {_version_tag}  ·  {_commit}  ·  {_date}', 'muted'),
                ]

                title_font   = Font(bold=True, size=18, color='1D1D1F')
                subtitle_fnt = Font(size=11, color='6E6E73')
                heading_font = Font(bold=True, size=11, color='4C8BF5')
                body_font    = Font(size=11, color='1D1D1F')
                name_font    = Font(bold=True, size=14, color='1D1D1F')
                muted_font   = Font(size=9, color='AEAEB2')
                cell_align   = Alignment(wrap_text=True, vertical='top')

                font_map = {
                    'title':    title_font,
                    'subtitle': subtitle_fnt,
                    'heading':  heading_font,
                    'body':     body_font,
                    'name':     name_font,
                    'muted':    muted_font,
                    None:       body_font,
                }

                for r_idx, (text, style_key) in enumerate(about_rows, start=1):
                    cell = ws_about.cell(row=r_idx, column=1, value=text)
                    cell.font = font_map.get(style_key, body_font)
                    cell.alignment = cell_align

                ws_about.column_dimensions['A'].width = 72
                # Row heights: title tall, headings medium, body auto via wrap
                ws_about.row_dimensions[1].height = 30   # title
                ws_about.row_dimensions[2].height = 18   # subtitle
                for r_idx, (_, style_key) in enumerate(about_rows, start=1):
                    if style_key == 'body':
                        ws_about.row_dimensions[r_idx].height = 42

            messagebox.showinfo("Export Successful", f"Mapping file saved to:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export mapping file:\n{str(e)}")

    def smart_guess_excel_column(self, pdf_field_name: str) -> str:
        """Generate a smart guess for Excel column name."""
        # Convert underscores to spaces
        name = pdf_field_name.replace('_', ' ')

        # Known mappings
        mappings = {
            'first name': 'First name',
            'surname': 'surname',
            'vcaa number': 'VCAA student number',
            'vcaa student number': 'VCAA student number',
            'school name': 'School name',
            'vcaa school code': 'VCAA school code',
        }

        lower_name = name.lower()
        if lower_name in mappings:
            return mappings[lower_name]

        return name

    def generate_field_notes(self, field: PDFField) -> str:
        """Generate helpful notes for a field."""
        notes = []

        if field.field_name.lower() in ['surname', 'first_name', 'first name']:
            notes.append('Used for filename')

        if field.field_name.lower() in ['surname', 'first name', 'vcaa student number', 'vcaa_number']:
            notes.append('Critical field')

        if field.is_combed and 'dob' in field.field_name.lower():
            notes.append('Auto-format date')

        return ', '.join(notes) if notes else ''

    def save_template_config(self):
        """Save template configuration to JSON."""
        if not self.analyzed_fields:
            messagebox.showwarning("No Data", "Please analyze a PDF template first.")
            return

        template_name = self.template_name_var.get()
        if not template_name:
            messagebox.showwarning("No Name", "Please provide a template name.")
            return

        try:
            # Create config
            field_stats = {}
            for field in self.analyzed_fields:
                key = field.field_type.lower().replace('-', '_').replace(' ', '_')
                field_stats[key] = field_stats.get(key, 0) + 1

            # Collect per-field data types for persistence (save all types
            # so that explicit 'text' choices override smart-guess defaults)
            field_data_types = {f.field_name: f.data_type for f in self.analyzed_fields}

            # Collect field type overrides (only save single-field combed overrides)
            field_type_overrides = {}
            for f in self.analyzed_fields:
                if f.is_combed and not f.combed_fields:
                    # Single-field combed (user-marked or MaxLen-detected)
                    field_type_overrides[f.field_name] = {
                        'field_type': f.field_type,
                        'length': f.length,
                    }

            # Collect explicit field→Excel column mappings set in Tab 2
            field_excel_columns = {
                f.field_name: f.excel_column
                for f in self.analyzed_fields
                if f.excel_column is not None
            }

            config = TemplateConfig(
                template_name=template_name,
                pdf_filename=os.path.basename(self.pdf_template_path.get()),
                pdf_path=self.pdf_template_path.get(),
                created_date=datetime.now().isoformat(),
                last_used=datetime.now().isoformat(),
                total_fields=len(self.analyzed_fields),
                field_types=field_stats,
                mapping_file=f"{template_name}_mapping.xlsx",
                use_auto_matching=True,
                critical_fields=['surname', 'First name', 'VCAA student number'],
                field_data_types=field_data_types,
                field_type_overrides=field_type_overrides,
                field_excel_columns=field_excel_columns,
                notes="",
                version="2.0"
            )

            # Save to file
            config_path = os.path.join(self.settings.templates_directory, f"{template_name}.json")
            config.save_to_file(config_path)

            self.current_template = config

            messagebox.showinfo("Saved", f"Template configuration saved to:\n{config_path}")

            # Refresh recent templates
            self.populate_recent_templates()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template config:\n{str(e)}")

    # ========== TAB 2: MAP FIELDS ==========

    def setup_tab2_mapping(self):
        """Setup Tab 2: Field Mapping — live editor for PDF field → Excel column mappings."""
        tab = self.tab2
        C = COLORS

        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        # ── Card 1: Data File ──────────────────────────────────────────
        file_inner = self.create_section(
            container,
            "Data File",
            subtitle="The Excel or CSV file whose columns will be mapped to PDF fields.",
        )

        self._tab2_file_label = tk.Label(
            file_inner,
            text="No data file loaded — load your data file on the Generate tab first.",
            font=font(10),
            fg=C['text_tertiary'],
            bg=C['bg_surface'],
            anchor='w',
            wraplength=680,
        )
        self._tab2_file_label.pack(anchor='w', pady=(0, 4))

        # ── Card 2: Field Mappings ─────────────────────────────────────
        mappings_inner = self.create_section(
            container,
            "Field Mappings",
            subtitle="Match each PDF form field to the Excel column that contains its data.",
            expand=True,
        )

        # Status line (updated by _update_mapping_status)
        self._tab2_status_label = tk.Label(
            mappings_inner,
            text="Analyze a PDF template on Tab 1 to begin.",
            font=font(10),
            fg=C['text_secondary'],
            bg=C['bg_surface'],
            anchor='w',
        )
        self._tab2_status_label.pack(anchor='w', pady=(0, SPACING['element_gap']))

        # Column header row
        header_frame = tk.Frame(mappings_inner, bg=C['bg_surface'])
        header_frame.pack(fill=tk.X, pady=(0, 4))
        tk.Label(header_frame, text="PDF Field", font=font(9, 'bold'),
                 fg=C['text_secondary'], bg=C['bg_surface'], width=28, anchor='w').pack(side=tk.LEFT)
        tk.Label(header_frame, text="Excel Column", font=font(9, 'bold'),
                 fg=C['text_secondary'], bg=C['bg_surface'], width=32, anchor='w').pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(header_frame, text="Status", font=font(9, 'bold'),
                 fg=C['text_secondary'], bg=C['bg_surface'], width=10, anchor='w').pack(side=tk.LEFT, padx=(8, 0))

        # Thin separator
        tk.Frame(mappings_inner, bg=C['border_subtle'], height=1).pack(fill=tk.X, pady=(0, 6))

        # Scrollable frame that holds the per-field rows
        scroll = ScrollableFrame(mappings_inner)
        scroll.pack(fill=tk.BOTH, expand=True)
        self._tab2_mapping_frame = scroll.scrollable_frame

        # ── Buttons ────────────────────────────────────────────────────
        btn_frame = tk.Frame(container, bg=C['bg_base'])
        btn_frame.pack(fill=tk.X, pady=(SPACING['element_gap'], 0))

        self._tab2_auto_btn = ttk.Button(
            btn_frame,
            text="Auto-Map All",
            command=self._auto_map_fields,
            state=tk.DISABLED,
        )
        self._tab2_auto_btn.pack(side=tk.LEFT, padx=(0, 8))

        self._tab2_clear_btn = ttk.Button(
            btn_frame,
            text="Clear All Mappings",
            command=self._clear_all_mappings,
            style='Secondary.TButton',
            state=tk.DISABLED,
        )
        self._tab2_clear_btn.pack(side=tk.LEFT)

    # ── Tab 2 helper methods ───────────────────────────────────────────────

    def _refresh_tab2_mappings(self):
        """Rebuild the Tab 2 mapping rows from current analyzed_fields and df columns.

        Safe to call at any time; handles all partial-state combinations gracefully.
        """
        # Guard: widgets may not exist yet if called before setup_ui completes
        if self._tab2_status_label is None:
            return

        C = COLORS

        # Update the data-file label
        excel_path = self.excel_file_path.get()
        if excel_path:
            self._tab2_file_label.config(
                text=os.path.basename(excel_path),
                fg=C['text_primary'],
            )
        else:
            self._tab2_file_label.config(
                text="No data file loaded — load your data file on the Generate tab first.",
                fg=C['text_tertiary'],
            )

        # Case 1: No PDF analyzed yet
        if not self.analyzed_fields:
            self._tab2_status_label.config(
                text="Analyze a PDF template on Tab 1 to begin.",
                fg=C['text_secondary'],
            )
            self._clear_mapping_rows()
            self._set_mapping_buttons(enabled=False)
            return

        # Case 2: PDF analyzed but no Excel loaded
        if self.df is None:
            self._tab2_status_label.config(
                text="Load your data file on the Generate tab to see available columns.",
                fg=C['text_secondary'],
            )
            self._clear_mapping_rows()
            self._set_mapping_buttons(enabled=False)
            return

        # Case 3: Both available — build the mapping rows
        self._clear_mapping_rows()
        self._tab2_combos.clear()

        col_list = sorted(self.df.columns.tolist())
        column_options = ["-- not mapped --"] + col_list
        col_set = set(self.df.columns.tolist())

        for field in self.analyzed_fields:
            row = tk.Frame(self._tab2_mapping_frame, bg=C['bg_surface'])
            row.pack(fill=tk.X, pady=2)

            # PDF field name label
            tk.Label(
                row,
                text=field.field_name,
                font=font(10),
                fg=C['text_primary'],
                bg=C['bg_surface'],
                width=28,
                anchor='w',
            ).pack(side=tk.LEFT)

            # Determine initial combobox value
            if field.excel_column and field.excel_column in col_set:
                initial = field.excel_column
            else:
                # Saved column no longer in file — clear it
                field.excel_column = None
                initial = "-- not mapped --"

            combo = ttk.Combobox(
                row,
                values=column_options,
                state="readonly",
                width=32,
            )
            combo.set(initial)
            combo.pack(side=tk.LEFT, padx=(8, 0))

            # Status icon label
            mapped = (initial != "-- not mapped --")
            status_lbl = tk.Label(
                row,
                text="✓" if mapped else "–",
                font=font(10),
                fg=C['success'] if mapped else C['text_tertiary'],
                bg=C['bg_surface'],
                width=4,
                anchor='w',
            )
            status_lbl.pack(side=tk.LEFT, padx=(8, 0))

            self._tab2_combos[field.field_name] = (combo, status_lbl)

            # Bind selection event (use default arg to capture loop variable)
            combo.bind(
                '<<ComboboxSelected>>',
                lambda e, fn=field.field_name, cb=combo, sl=status_lbl:
                    self._on_mapping_changed(fn, cb, sl),
            )

        self._update_mapping_status()
        self._set_mapping_buttons(enabled=True)

    def _clear_mapping_rows(self):
        """Destroy all child widgets inside the mapping scroll frame."""
        if self._tab2_mapping_frame is None:
            return
        for widget in self._tab2_mapping_frame.winfo_children():
            widget.destroy()

    def _set_mapping_buttons(self, enabled: bool):
        """Enable or disable the Auto-Map and Clear All buttons."""
        state = tk.NORMAL if enabled else tk.DISABLED
        if self._tab2_auto_btn:
            self._tab2_auto_btn.config(state=state)
        if self._tab2_clear_btn:
            self._tab2_clear_btn.config(state=state)

    def _on_mapping_changed(self, field_name: str, combo: ttk.Combobox, status_lbl: tk.Label):
        """Handle user changing a field mapping dropdown."""
        val = combo.get()
        for field in self.analyzed_fields:
            if field.field_name == field_name:
                field.excel_column = None if val == "-- not mapped --" else val
                break
        mapped = (val != "-- not mapped --")
        status_lbl.config(
            text="✓" if mapped else "–",
            fg=COLORS['success'] if mapped else COLORS['text_tertiary'],
        )
        self._update_mapping_status()

    def _auto_map_fields(self):
        """Apply smart-guess mappings to all fields, overwriting existing mappings."""
        if not self.df or not self.analyzed_fields:
            return
        column_lower = {col.lower(): col for col in self.df.columns}
        for field in self.analyzed_fields:
            # Try smart-guess name first, then underscore-stripped field name, then direct
            guess = self.smart_guess_excel_column(field.field_name)
            if guess.lower() in column_lower:
                field.excel_column = column_lower[guess.lower()]
            elif field.field_name.replace('_', ' ').lower() in column_lower:
                field.excel_column = column_lower[field.field_name.replace('_', ' ').lower()]
            elif field.field_name.lower() in column_lower:
                field.excel_column = column_lower[field.field_name.lower()]
            # else: no match found, leave as-is
        self._refresh_tab2_mappings()

    def _clear_all_mappings(self):
        """Reset all field→Excel column mappings to unset."""
        for field in self.analyzed_fields:
            field.excel_column = None
        self._refresh_tab2_mappings()

    def _update_mapping_status(self):
        """Update the 'X of Y fields mapped' status label in Tab 2."""
        if self._tab2_status_label is None or not self.analyzed_fields:
            return
        total = len(self.analyzed_fields)
        mapped = sum(1 for f in self.analyzed_fields if f.excel_column is not None)
        C = COLORS
        if mapped == total:
            self._tab2_status_label.config(
                text=f"All {total} fields mapped \u2713",
                fg=C['success'],
            )
        elif mapped == 0:
            self._tab2_status_label.config(
                text="No fields mapped \u2014 use Auto-Map or set manually.",
                fg=C['text_secondary'],
            )
        else:
            unmatched = total - mapped
            self._tab2_status_label.config(
                text=f"{mapped} of {total} fields mapped, {unmatched} unmatched.",
                fg=C['warning'],
            )

    # ========== TAB 3: GENERATE PDFs (From original app) ==========

    def setup_tab3_generate(self):
        """Setup Tab 3: PDF Generation (original functionality)."""
        tab = self.tab3

        container = ttk.Frame(tab, padding=str(SPACING['page_padding']))
        container.pack(fill=tk.BOTH, expand=True)

        # Template selection bar
        template_frame = tk.Frame(container, bg=COLORS['bg_base'])
        template_frame.pack(fill=tk.X, pady=(0, SPACING['section_gap']))

        tk.Label(template_frame, text="Template:", font=font(10),
                 fg=COLORS['text_secondary'], bg=COLORS['bg_base']).pack(side=tk.LEFT, padx=(0, 10))

        self.selected_template_var = tk.StringVar()
        template_combo = ttk.Combobox(template_frame, textvariable=self.selected_template_var, state="readonly", width=40)
        template_combo.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(template_frame, text="Change Template", command=self.change_template_tab3).pack(side=tk.LEFT, padx=(0, 8))

        self.matching_status_label = ttk.Label(template_frame, text="Auto-matching enabled", style='Success.TLabel')
        self.matching_status_label.pack(side=tk.LEFT, padx=10)

        # File Selection section
        file_inner = self.create_section(container, "Select Files")

        # PDF Template selection
        pdf_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        pdf_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(pdf_row, text="PDF Template:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(pdf_row, textvariable=self.pdf_template_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(pdf_row, text="Browse...", command=self.select_pdf_tab3, width=10).pack(side=tk.LEFT)

        # Excel file selection
        excel_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        excel_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(excel_row, text="Excel Data File:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(excel_row, textvariable=self.excel_file_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(excel_row, text="Browse...", command=self.select_excel_tab3, width=10).pack(side=tk.LEFT)

        # Output folder selection
        output_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        output_row.pack(fill=tk.X, pady=(0, SPACING['element_gap']))
        tk.Label(output_row, text="Output Folder:", width=18, anchor=tk.W,
                 font=font(11), fg=COLORS['text_primary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT)
        ttk.Entry(output_row, textvariable=self.output_dir_path, width=50).pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        ttk.Button(output_row, text="Browse...", command=self.select_output_dir_tab3, width=10).pack(side=tk.LEFT)
        tk.Label(output_row, text="(optional)", font=font(9),
                 fg=COLORS['text_tertiary'], bg=COLORS['bg_surface']).pack(side=tk.LEFT, padx=(6, 0))

        # Load button
        btn_row = tk.Frame(file_inner, bg=COLORS['bg_surface'])
        btn_row.pack(fill=tk.X, pady=(SPACING['element_gap'], 0))
        ttk.Button(btn_row, text="Load & Preview Data", command=self.load_data_tab3, bootstyle='primary').pack()

        # Validation section
        validation_inner = self.create_section(container, "Validation")

        self.validation_text_tab3 = tk.Text(
            validation_inner, height=3, wrap=tk.WORD, state=tk.DISABLED,
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['text_primary'],
            selectbackground=COLORS['accent_subtle'],
            selectforeground=COLORS['text_primary'],
            relief='flat',
            borderwidth=0,
            padx=10,
            pady=8,
            font=font(10),
            autostyle=False,
        )
        self.validation_text_tab3.pack(fill=tk.X)
        # Keep a reference for pack_forget compatibility
        self.validation_frame_tab3 = validation_inner.master.master

        # Selection Controls
        selection_frame = tk.Frame(container, bg=COLORS['bg_base'])
        selection_frame.pack(fill=tk.X, pady=(0, SPACING['element_gap']))

        tk.Label(selection_frame, text="Select students to process:", font=font(11),
                 fg=COLORS['text_primary'], bg=COLORS['bg_base']).pack(side=tk.LEFT)

        ttk.Button(selection_frame, text="Select All", command=self.select_all_tab3).pack(side=tk.LEFT, padx=(15, 5))
        ttk.Button(selection_frame, text="Deselect All", command=self.deselect_all_tab3).pack(side=tk.LEFT, padx=(0, 5))

        self.selection_count_label_tab3 = ttk.Label(selection_frame, text="", style='Secondary.TLabel')
        self.selection_count_label_tab3.pack(side=tk.RIGHT)

        # Student Preview section
        preview_inner = self.create_section(container, "Student Preview",
                                            subtitle="Click rows to select or deselect", expand=True)

        # Treeview for student list
        columns = ('selected', 'row', 'surname', 'first_name', 'student_number', 'status')
        self.tree_tab3 = ttk.Treeview(preview_inner, columns=columns, show='headings', height=12)

        self.tree_tab3.heading('selected', text='Sel')
        self.tree_tab3.heading('row', text='#')
        self.tree_tab3.heading('surname', text='Surname')
        self.tree_tab3.heading('first_name', text='First Name')
        self.tree_tab3.heading('student_number', text='Student Number')
        self.tree_tab3.heading('status', text='Status')

        self.tree_tab3.column('selected', width=40, anchor='center')
        self.tree_tab3.column('row', width=40, anchor='center')
        self.tree_tab3.column('surname', width=140)
        self.tree_tab3.column('first_name', width=140)
        self.tree_tab3.column('student_number', width=110)
        self.tree_tab3.column('status', width=180)

        # Bind click event for toggling selection
        self.tree_tab3.bind('<ButtonRelease-1>', self.toggle_selection_tab3)

        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(preview_inner, orient=tk.VERTICAL, command=self.tree_tab3.yview)
        self.tree_tab3.configure(yscrollcommand=scrollbar.set)

        self.tree_tab3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Setup treeview tags and hover
        setup_treeview_tags(self.tree_tab3)
        bind_treeview_hover(self.tree_tab3)

        # Summary label
        self.summary_label_tab3 = ttk.Label(container, text="No data loaded", style='Secondary.TLabel')
        self.summary_label_tab3.pack(pady=(8, 4))

        # Progress Frame
        progress_frame = tk.Frame(container, bg=COLORS['bg_base'])
        progress_frame.pack(fill=tk.X, pady=(0, SPACING['section_gap']))

        self.progress_var_tab3 = tk.DoubleVar()
        self.progress_bar_tab3 = ttk.Progressbar(progress_frame, variable=self.progress_var_tab3, maximum=100)
        self.progress_bar_tab3.pack(fill=tk.X, pady=(4, 6))

        self.progress_label_tab3 = ttk.Label(progress_frame, text="", style='Secondary.TLabel')
        self.progress_label_tab3.pack()

        # Generate Button (large CTA)
        self.generate_btn_tab3 = ttk.Button(
            container,
            text="Generate PDFs for Selected Students",
            command=self.start_generation_tab3,
            state=tk.DISABLED,
            bootstyle='primary',
            padding=(32, 14),
        )
        self.generate_btn_tab3.pack(pady=SPACING['section_gap'])

    def select_pdf_tab3(self):
        """Select PDF in Tab 3."""
        filepath = filedialog.askopenfilename(
            title="Select PDF Template",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filepath:
            self.pdf_template_path.set(filepath)

    def select_excel_tab3(self):
        """Select Excel file in Tab 3."""
        filepath = filedialog.askopenfilename(
            title="Select Student Data Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            self.excel_file_path.set(filepath)

    def select_output_dir_tab3(self):
        """Select output directory for generated PDFs."""
        initial_dir = self.output_dir_path.get() or os.path.dirname(self.excel_file_path.get() or '')
        dirpath = filedialog.askdirectory(
            title="Select Output Folder for Generated PDFs",
            initialdir=initial_dir or None,
        )
        if dirpath:
            self.output_dir_path.set(dirpath)

    def change_template_tab3(self):
        """Change template in Tab 3."""
        self.load_template_from_file()

    def _pick_excel_sheet(self, sheet_names: list) -> Optional[str]:
        """Show a modal sheet-picker dialog and return the chosen sheet name.

        Returns None if the user cancels.
        """
        C = COLORS
        ff = SYSTEM_FONTS['family']
        result = {'choice': None}

        dialog = tk.Toplevel(self.root)
        dialog.title("Select Sheet")
        dialog.resizable(False, False)
        dialog.configure(bg=C['bg_base'])
        dialog.grab_set()

        # Centre over the main window
        dialog.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width() // 2) - 220
        py = self.root.winfo_y() + (self.root.winfo_height() // 2) - 80
        dialog.geometry(f"440x210+{px}+{py}")

        # Content
        inner = tk.Frame(dialog, bg=C['bg_base'], padx=24, pady=20)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner,
            text="This workbook has multiple sheets.",
            font=(ff, 12, 'bold'), fg=C['text_primary'], bg=C['bg_base'],
        ).pack(anchor=tk.W)
        tk.Label(inner,
            text="Please select the sheet that contains your data:",
            font=(ff, 10), fg=C['text_secondary'], bg=C['bg_base'],
        ).pack(anchor=tk.W, pady=(2, 10))

        combo = ttk.Combobox(inner, values=sheet_names, state='readonly',
                             font=(ff, 11), width=36)
        combo.current(0)
        combo.pack(anchor=tk.W)

        # Buttons
        btn_row = tk.Frame(inner, bg=C['bg_base'])
        btn_row.pack(anchor=tk.E, pady=(14, 0))

        def on_ok():
            result['choice'] = combo.get()
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        ttk.Button(btn_row, text="Cancel", command=on_cancel).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="Load this sheet", bootstyle='primary',
                   command=on_ok).pack(side=tk.LEFT)

        dialog.bind('<Return>', lambda e: on_ok())
        dialog.bind('<Escape>', lambda e: on_cancel())
        dialog.wait_window()
        return result['choice']

    def load_data_tab3(self):
        """Load Excel data for Tab 3 (original functionality)."""
        pdf_path = self.pdf_template_path.get()
        excel_path = self.excel_file_path.get()

        # Validate paths
        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "Please select a valid PDF template file.")
            return

        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Error", "Please select a valid Excel data file.")
            return

        try:
            # Load PDF and get field names
            reader = PdfReader(pdf_path)
            self.pdf_fields = []  # Always reset before re-reading
            fields = reader.get_fields()
            if fields:
                self.pdf_fields = list(fields.keys())
            else:
                messagebox.showerror("Error", "The PDF template has no fillable form fields.")
                return

            # Load Excel data (case-insensitive extension check)
            if excel_path.lower().endswith('.csv'):
                try:
                    self.df = pd.read_csv(excel_path, encoding='utf-8-sig')
                except UnicodeDecodeError:
                    self.df = pd.read_csv(excel_path, encoding='latin-1')
            else:
                xl = pd.ExcelFile(excel_path)
                sheet_names = xl.sheet_names
                if len(sheet_names) == 1:
                    chosen_sheet = sheet_names[0]
                else:
                    chosen_sheet = self._pick_excel_sheet(sheet_names)
                    if chosen_sheet is None:
                        return  # User cancelled the dialog
                self.df = xl.parse(chosen_sheet, dtype=str)

            # Clean column names (strip whitespace, lowercase for matching)
            self.df.columns = [str(col).strip() for col in self.df.columns]

            # Create lowercase mapping for field matching
            self.column_mapping = {col.lower(): col for col in self.df.columns}
            self.field_mapping = {field.lower(): field for field in self.pdf_fields}

            # Clear previous selections
            self.selected_rows = {}

            # Validate and show preview
            self.validate_data_tab3()
            self.show_preview_tab3()

            # Select all by default
            self.select_all_tab3()

            # Enable generate button
            self.generate_btn_tab3.config(state=tk.NORMAL)

            # Refresh Tab 2 mapping dropdowns with the new column list;
            # auto-map any fields that don't yet have an explicit mapping
            if not self.analyzed_fields or any(f.excel_column is None for f in self.analyzed_fields):
                self._auto_map_fields()
            self._refresh_tab2_mappings()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data:\n{str(e)}")

    def validate_data_tab3(self):
        """Validate data for Tab 3."""
        warnings = []

        # Check each row for critical fields
        for idx, row in self.df.iterrows():
            row_warnings = []
            row_dict = {str(col).lower(): val for col, val in row.items()}

            for field in self.critical_fields:
                val = row_dict.get(field, '')
                if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                    row_warnings.append(field)

            if row_warnings:
                surname = row_dict.get('surname', f'Row {idx+1}')
                if pd.isna(surname) or str(surname).strip() == '':
                    surname = f'Row {idx+1}'
                warnings.append(f"• {surname}: Missing {', '.join(row_warnings)}")

        # Display warnings
        self.validation_text_tab3.config(state=tk.NORMAL)
        self.validation_text_tab3.delete(1.0, tk.END)

        if warnings:
            warning_text = f"{len(warnings)} student(s) have missing critical fields:\n"
            warning_text += "\n".join(warnings[:10])
            if len(warnings) > 10:
                warning_text += f"\n... and {len(warnings) - 10} more"
            self.validation_text_tab3.insert(1.0, warning_text)
            self.validation_text_tab3.config(fg=COLORS['warning'])
        else:
            self.validation_text_tab3.insert(1.0, "All students have required fields populated.")
            self.validation_text_tab3.config(fg=COLORS['success'])

        # Warn about PDF fields that have no explicit mapping and won't auto-match any Excel column
        if self.analyzed_fields and self.df is not None:
            col_names_lower = {str(col).lower() for col in self.df.columns}
            silent_blanks = [
                f for f in self.analyzed_fields
                if f.excel_column is None
                and f.field_name.lower() not in col_names_lower
            ]
            if silent_blanks:
                names = ", ".join(f.field_name for f in silent_blanks[:5])
                if len(silent_blanks) > 5:
                    names += f" (+{len(silent_blanks) - 5} more)"
                mapping_note = (
                    f"\n\n⚠ {len(silent_blanks)} field(s) have no mapping and won't "
                    f"auto-match — they will be blank:\n{names}\n"
                    f"Go to Tab 2 to set explicit mappings."
                )
                self.validation_text_tab3.config(state=tk.NORMAL)
                self.validation_text_tab3.insert(tk.END, mapping_note)
                self.validation_text_tab3.config(fg=COLORS['warning'])

        self.validation_text_tab3.config(state=tk.DISABLED)

    def show_preview_tab3(self):
        """Show preview for Tab 3."""
        # Clear existing items
        for item in self.tree_tab3.get_children():
            self.tree_tab3.delete(item)

        # Reset selected rows tracking
        self.selected_rows = {}

        # Add students to preview
        valid_count = 0
        warning_count = 0

        for idx, row in self.df.iterrows():
            row_dict = {str(col).lower(): val for col, val in row.items()}

            surname = str(row_dict.get('surname', '')).strip()
            first_name = str(row_dict.get('first name', '')).strip()
            student_num = str(row_dict.get('vcaa student number', '')).strip()

            # Skip completely empty rows
            if (pd.isna(row_dict.get('surname')) or surname == '' or surname.lower() == 'nan'):
                continue

            # Check status
            missing = []
            for field in self.critical_fields:
                val = row_dict.get(field, '')
                if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                    missing.append(field)

            if missing:
                status = f"Missing: {', '.join(missing)}"
                warning_count += 1
            else:
                status = "Ready"
                valid_count += 1

            # Clean display values
            if surname.lower() == 'nan':
                surname = ''
            if first_name.lower() == 'nan':
                first_name = ''
            if student_num.lower() == 'nan':
                student_num = ''

            # Build tags: alternating rows + status
            row_num = valid_count + warning_count
            row_tag = 'odd' if row_num % 2 else 'even'
            status_tag = 'warning_row' if missing else 'success_row'

            # Insert with empty checkbox initially
            item_id = self.tree_tab3.insert('', tk.END,
                values=('', idx+1, surname, first_name, student_num, status),
                tags=(row_tag, status_tag),
            )
            # Store mapping of item_id to dataframe index
            self.selected_rows[item_id] = {'index': idx, 'selected': False}

        # Update summary
        total = valid_count + warning_count
        self.summary_label_tab3.config(
            text=f"Loaded {total} students: {valid_count} ready, {warning_count} with warnings"
        )

        self.update_selection_count_tab3()

    def toggle_selection_tab3(self, event):
        """Toggle selection in Tab 3."""
        item = self.tree_tab3.identify_row(event.y)
        if item and item in self.selected_rows:
            # Toggle the selection
            self.selected_rows[item]['selected'] = not self.selected_rows[item]['selected']

            # Update the checkbox display
            current_values = list(self.tree_tab3.item(item, 'values'))
            current_values[0] = 'Yes' if self.selected_rows[item]['selected'] else ''
            self.tree_tab3.item(item, values=current_values)

            self.update_selection_count_tab3()

    def select_all_tab3(self):
        """Select all in Tab 3."""
        for item_id in self.selected_rows:
            self.selected_rows[item_id]['selected'] = True
            current_values = list(self.tree_tab3.item(item_id, 'values'))
            current_values[0] = 'Yes'
            self.tree_tab3.item(item_id, values=current_values)

        self.update_selection_count_tab3()

    def deselect_all_tab3(self):
        """Deselect all in Tab 3."""
        for item_id in self.selected_rows:
            self.selected_rows[item_id]['selected'] = False
            current_values = list(self.tree_tab3.item(item_id, 'values'))
            current_values[0] = ''
            self.tree_tab3.item(item_id, values=current_values)

        self.update_selection_count_tab3()

    def update_selection_count_tab3(self):
        """Update selection count for Tab 3."""
        selected = sum(1 for item in self.selected_rows.values() if item['selected'])
        total = len(self.selected_rows)
        self.selection_count_label_tab3.config(text=f"Selected: {selected} of {total}")

        # Update button text
        if selected == 0:
            self.generate_btn_tab3.config(text="Generate PDFs (none selected)", state=tk.DISABLED)
        elif selected == 1:
            self.generate_btn_tab3.config(text="Generate PDF for 1 Student", state=tk.NORMAL)
        else:
            self.generate_btn_tab3.config(text=f"Generate PDFs for {selected} Students", state=tk.NORMAL)

    def start_generation_tab3(self):
        """Start PDF generation for Tab 3.

        All shared state is snapshot here (main thread) so the worker
        thread never touches tkinter StringVars or mutable instance data.
        """
        # Check if any students are selected
        selected_count = sum(1 for item in self.selected_rows.values() if item['selected'])
        if selected_count == 0:
            messagebox.showwarning("No Selection", "Please select at least one student to generate PDFs.")
            return

        # ── Snapshot all shared state in the main thread ──
        ctx = {
            'pdf_path': self.pdf_template_path.get(),
            'excel_path': self.excel_file_path.get(),
            'df': self.df.copy(),
            'selected_indices': [
                item['index'] for item in self.selected_rows.values()
                if item['selected']
            ],
            'analyzed_fields': copy.deepcopy(self.analyzed_fields),
            'pdf_fields': list(self.pdf_fields) if hasattr(self, 'pdf_fields') else [],
            'school_name': self.settings.school_name or "School",
            'school_year': self.settings.school_year or str(datetime.now().year),
            'combed_padding': self.settings.combed_field_padding,
            'combed_align': self.settings.combed_field_align,
            'output_dir': self.output_dir_path.get().strip() or '',
        }

        # Reset progress bar and disable button
        self.progress_var_tab3.set(0)
        self.generate_btn_tab3.config(state=tk.DISABLED)
        self.update_status(f"Generating PDFs for {selected_count} students...", 'info')

        # Start generation in background thread with snapshot
        thread = threading.Thread(target=self.run_generation_tab3, args=(ctx,))
        thread.daemon = True
        thread.start()

    def run_generation_tab3(self, ctx):
        """Run PDF generation in a background thread.

        Uses only the snapshot *ctx* dict — never reads tkinter
        StringVars or mutable instance state directly.
        """
        try:
            # Determine output folder: use custom dir if set, else default
            if ctx['output_dir'] and os.path.isdir(ctx['output_dir']):
                output_folder = ctx['output_dir']
            else:
                excel_dir = os.path.dirname(ctx['excel_path'])
                output_folder = os.path.join(excel_dir, "Completed Applications")
            os.makedirs(output_folder, exist_ok=True)

            total = len(ctx['selected_indices'])
            success_count = 0
            error_count = 0

            for i, idx in enumerate(ctx['selected_indices']):
                row = ctx['df'].iloc[idx]
                row_dict = {str(col).lower(): val for col, val in row.items()}

                # Get name for filename
                first_name = str(row_dict.get('first name', 'Unknown')).strip()
                surname = str(row_dict.get('surname', 'Unknown')).strip()

                if first_name.lower() == 'nan':
                    first_name = 'Unknown'
                if surname.lower() == 'nan':
                    surname = 'Unknown'

                # Clean names for filename (remove invalid characters)
                safe_first = "".join(c for c in first_name if c.isalnum() or c in ' -_').strip()
                safe_surname = "".join(c for c in surname if c.isalnum() or c in ' -_').strip()
                safe_school = "".join(c for c in ctx['school_name'] if c.isalnum() or c in ' -_').strip()
                safe_year = "".join(c for c in ctx['school_year'] if c.isalnum() or c in ' -_').strip()

                # Create filename
                filename = f"{safe_first}_{safe_surname}_Evidence Application {safe_school} {safe_year}.pdf"
                output_path = os.path.join(output_folder, filename)

                # Avoid overwriting existing files (e.g. duplicate names, re-runs)
                if os.path.exists(output_path):
                    base, ext = os.path.splitext(output_path)
                    counter = 1
                    while os.path.exists(f"{base} ({counter}){ext}"):
                        counter += 1
                    output_path = f"{base} ({counter}){ext}"

                try:
                    self._generate_single_pdf(ctx, row, output_path)
                    success_count += 1
                    status_text = f"Created: {filename}"
                except Exception as e:
                    error_count += 1
                    status_text = f"Error: {surname} - {str(e)}"

                # Update progress (UI calls are safe via root.after)
                progress = ((i + 1) / total) * 100
                self.root.after(0, self.update_progress_tab3, progress, status_text, i+1, total)

            # Final message
            final_message = f"Complete! {success_count} PDFs created"
            if error_count > 0:
                final_message += f", {error_count} errors"
            final_message += f"\n\nOutput folder: {output_folder}"

            self.root.after(0, self.generation_complete_tab3, final_message, output_folder)

        except Exception as e:
            err_msg = str(e)  # Capture before 'e' goes out of scope (PEP 3110)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Error", f"Generation failed:\n{msg}"))
            self.root.after(0, lambda: self.generate_btn_tab3.config(state=tk.NORMAL))

    def _generate_single_pdf(self, ctx, row_data, output_path):
        """Generate a single PDF with combed field support.

        THREAD SAFETY: This method runs on the generation worker thread.
        It must only access local variables and the immutable *ctx* snapshot
        dict — never read or write any ``self.*`` attribute directly.
        All UI updates must be dispatched via ``self.root.after()``.
        """
        reader = PdfReader(ctx['pdf_path'])
        writer = PdfWriter()

        # Clone the PDF
        writer.append(reader)

        # Create a dictionary of field values to fill
        field_values = {}

        # Check if we have analyzed fields with combed field metadata
        if ctx['analyzed_fields']:
            # Use combed field filler for smart filling
            combed_filler = CombedFieldFiller(settings={
                'padding': ctx['combed_padding'],
                'align': ctx['combed_align']
            })

            # Build lookup of raw values keyed by lowercase column name
            row_raw_lower = {str(col).lower(): val for col, val in row_data.items()}

            # Build lookup of field data types
            field_types_lookup = {f.field_name.lower(): f.data_type
                                  for f in ctx['analyzed_fields']}

            # Process each analyzed field
            for field in ctx['analyzed_fields']:
                # Find matching Excel column:
                # 1. Explicit mapping from Tab 2 (field.excel_column) takes priority
                # 2. Fall back to auto-match by PDF field name (case-insensitive)
                if field.excel_column:
                    raw_val = row_raw_lower.get(field.excel_column.lower())
                else:
                    raw_val = row_raw_lower.get(field.field_name.lower())

                if raw_val is None:
                    continue

                # Format with type awareness
                value = self.format_value_tab3(raw_val, data_type=field.data_type)
                if not value:
                    continue

                # Fill field (handles both combed and regular)
                filled_values = combed_filler.fill_field(field, value)
                field_values.update(filled_values)

        else:
            # Fallback to original auto-matching (no combed field support)
            row_dict_lower = {str(col).lower(): val for col, val in row_data.items()}

            for pdf_field in ctx['pdf_fields']:
                pdf_field_lower = pdf_field.lower()

                # Try to find matching Excel column
                if pdf_field_lower in row_dict_lower:
                    val = self.format_value_tab3(row_dict_lower[pdf_field_lower])
                    field_values[pdf_field] = val

        # Split values into regular fields and single-field comb fields.
        # Comb fields need auto_regenerate=True so pypdf builds the
        # per-character appearance stream; regular fields use False to
        # avoid a spurious /NeedAppearances flag in some viewers.
        comb_field_names = set()
        if ctx['analyzed_fields']:
            for f in ctx['analyzed_fields']:
                if f.is_combed and not f.combed_fields:
                    comb_field_names.add(f.field_name)

        regular_values = {k: v for k, v in field_values.items()
                          if k not in comb_field_names}
        comb_values = {k: v for k, v in field_values.items()
                       if k in comb_field_names}

        for page in writer.pages:
            if regular_values:
                writer.update_page_form_field_values(
                    page, regular_values, auto_regenerate=False
                )
            if comb_values:
                writer.update_page_form_field_values(
                    page, comb_values, auto_regenerate=True
                )

        # Save the filled PDF
        with open(output_path, 'wb') as f:
            writer.write(f)

    @staticmethod
    def format_value_tab3(val, data_type: str = "text"):
        """Format a value for PDF insertion.

        Args:
            val: The raw cell value from pandas.
            data_type: One of 'text', 'number', or 'date'.
                       When 'date', Excel serial numbers are converted
                       to Australian DD/MM/YYYY format.
        """
        if pd.isna(val):
            return ""

        # Handle datetime objects (Australian format) regardless of data_type
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.strftime('%d/%m/%Y')

        # Date type: handle pandas-stringified datetimes e.g. "2024-05-01 00:00:00"
        # (occurs when dtype=str is used on read_excel — Timestamps become strings)
        if data_type == "date" and isinstance(val, str):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                try:
                    return datetime.strptime(val.strip(), fmt).strftime('%d/%m/%Y')
                except ValueError:
                    pass

        # Date type: convert Excel serial numbers to DD/MM/YYYY
        if data_type == "date" and isinstance(val, (int, float)):
            serial = int(val)
            # Valid Excel date serials: 1 (1900-01-01) to 2958465 (9999-12-31)
            if 1 <= serial <= 2958465:
                try:
                    from datetime import timedelta
                    # Excel epoch is 1899-12-30 (accounts for the 1900 leap-year bug)
                    excel_epoch = datetime(1899, 12, 30)
                    date_obj = excel_epoch + timedelta(days=serial)
                    return date_obj.strftime('%d/%m/%Y')
                except (ValueError, OverflowError):
                    pass  # Fall through to string conversion

        # Number type: strip trailing .0 from whole numbers
        if data_type == "number" and isinstance(val, float) and val == int(val):
            return str(int(val))

        # Convert to string and clean
        str_val = str(val).strip()
        if str_val.lower() == 'nan':
            return ""

        return str_val

    def update_progress_tab3(self, progress, status, current, total):
        """Update progress for Tab 3."""
        self.progress_var_tab3.set(progress)
        self.progress_label_tab3.config(text=f"[{current}/{total}] {status}")

    def generation_complete_tab3(self, message, output_folder):
        """Handle generation completion for Tab 3."""
        self.progress_label_tab3.config(text="Generation complete!")
        self.update_status("Generation complete!", 'success')
        self.update_selection_count_tab3()  # Re-enable button with correct state

        # Ask to open folder
        result = messagebox.askyesno(
            "Generation Complete",
            f"{message}\n\nWould you like to open the output folder?"
        )

        if result:
            # Open folder in system file manager
            try:
                if sys.platform == 'darwin':
                    import subprocess
                    subprocess.run(['open', output_folder], check=False)
                elif sys.platform == 'win32':
                    os.startfile(output_folder)
                else:
                    import subprocess
                    subprocess.run(['xdg-open', output_folder], check=False)
            except Exception:
                pass  # Failing to open folder is non-critical


def main():
    # On Windows, set the AppUserModelID BEFORE creating any windows.
    # This tells the Windows taskbar to use the .exe's embedded icon
    # instead of the default tkinter/ttkbootstrap feather icon.
    if sys.platform == 'win32':
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            'Antigravity.BulkPDFGenerator.v2'
        )

    # Must be called before Tk window creation for best DPI results
    from ttkbootstrap.utility import enable_high_dpi_awareness
    enable_high_dpi_awareness()

    root = tk.Tk()

    # Force window to front on launch
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))

    # Apply ttkbootstrap theme + custom styles
    apply_dark_theme(root)

    # Set window icon AFTER ttkbootstrap theme (Style() can reset icons).
    # iconphoto with a PNG is more reliable than iconbitmap on Windows
    # for controlling the taskbar icon when ttkbootstrap is in use.
    ico_path = get_resource_path('icon.ico')
    png_path = get_resource_path('icon.png')
    if sys.platform == 'win32':
        if os.path.exists(ico_path):
            root.iconbitmap(default=ico_path)
            root.iconbitmap(ico_path)
        if os.path.exists(png_path):
            _img = Image.open(png_path).convert('RGBA')
            _icon = ImageTk.PhotoImage(_img.resize((256, 256), Image.LANCZOS))
            root.iconphoto(True, _icon)
            root._icon_ref = _icon  # prevent garbage collection
    elif os.path.exists(png_path):
        _img = Image.open(png_path).convert('RGBA')
        _icon = ImageTk.PhotoImage(_img.resize((64, 64), Image.LANCZOS))
        root.iconphoto(True, _icon)
        root._icon_ref = _icon

    app = BulkPDFGenerator(root)
    root.mainloop()


if __name__ == "__main__":
    main()
